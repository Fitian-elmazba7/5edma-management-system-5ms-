from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QTableWidget, QTableWidgetItem, QHeaderView,
                             QComboBox, QGroupBox, QPushButton, QMessageBox,
                             QSplitter, QDateEdit, QProgressBar, QTabWidget,
                             QFrame, QCheckBox, QSpinBox, QLineEdit,
                             QCalendarWidget, QDialog, QDialogButtonBox,
                             QFormLayout, QRadioButton, QButtonGroup, QScrollArea)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QIcon, QFont, QColor
from utils.database import DatabaseManager
from utils.excel_handler import ExcelHandler
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class DateRangeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("اختر الفترة الزمنية")
        self.setModal(True)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()
        
        # اختيار تاريخ البداية
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-7))
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat("yyyy-MM-dd")
        
        # اختيار تاريخ النهاية
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat("yyyy-MM-dd")
        
        # خيارات الفترة المحددة
        self.period_group = QButtonGroup()
        
        self.week_radio = QRadioButton("أسبوع واحد (7 أيام)")
        self.week_radio.setChecked(True)
        self.week_radio.toggled.connect(lambda: self.set_period_days(7))
        
        self.two_weeks_radio = QRadioButton("أسبوعين (14 يوم)")
        self.two_weeks_radio.toggled.connect(lambda: self.set_period_days(14))
        
        self.month_radio = QRadioButton("شهر (30 يوم)")
        self.month_radio.toggled.connect(lambda: self.set_period_days(30))
        
        self.custom_radio = QRadioButton("مخصص")
        
        self.period_group.addButton(self.week_radio)
        self.period_group.addButton(self.two_weeks_radio)
        self.period_group.addButton(self.month_radio)
        self.period_group.addButton(self.custom_radio)
        
        # خيار أيام الخدمة فقط
        self.service_days_checkbox = QCheckBox("أيام الخدمة فقط (الخميس)")
        self.service_days_checkbox.setChecked(False)
        
        # خيار إضافة أيام إضافية للخدمة
        self.additional_days_group = QGroupBox("أيام خدمة إضافية")
        additional_days_layout = QVBoxLayout()
        
        self.monday_check = QCheckBox("الإثنين")
        self.tuesday_check = QCheckBox("الثلاثاء")
        self.wednesday_check = QCheckBox("الأربعاء")
        self.friday_check = QCheckBox("الجمعة")
        self.saturday_check = QCheckBox("السبت")
        self.sunday_check = QCheckBox("الأحد")
        
        additional_days_layout.addWidget(self.monday_check)
        additional_days_layout.addWidget(self.tuesday_check)
        additional_days_layout.addWidget(self.wednesday_check)
        additional_days_layout.addWidget(self.friday_check)
        additional_days_layout.addWidget(self.saturday_check)
        additional_days_layout.addWidget(self.sunday_check)
        
        self.additional_days_group.setLayout(additional_days_layout)
        self.additional_days_group.setVisible(False)
        
        # ربط حدث تغيير خيار أيام الخدمة
        self.service_days_checkbox.stateChanged.connect(self.toggle_additional_days)
        
        form_layout.addRow("تاريخ البداية:", self.start_date)
        form_layout.addRow("تاريخ النهاية:", self.end_date)
        form_layout.addRow("", self.week_radio)
        form_layout.addRow("", self.two_weeks_radio)
        form_layout.addRow("", self.month_radio)
        form_layout.addRow("", self.custom_radio)
        form_layout.addRow("", self.service_days_checkbox)
        form_layout.addRow("", self.additional_days_group)
        
        layout.addLayout(form_layout)
        
        # أزرار
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
    
    def toggle_additional_days(self, state):
        """إظهار/إخفاء أيام الخدمة الإضافية"""
        self.additional_days_group.setVisible(state == Qt.Checked)
    
    def set_period_days(self, days):
        if not self.sender().isChecked():
            return
        self.end_date.setDate(QDate.currentDate())
        self.start_date.setDate(QDate.currentDate().addDays(-days))
    
    def get_selected_service_days(self):
        """الحصول على أيام الخدمة المختارة"""
        service_days = ['Thursday']  # الخميس كأساسي
        
        # تحويل أيام الأسبوع المختارة إلى الإنجليزية
        day_mapping = {
            'الإثنين': 'Monday',
            'الثلاثاء': 'Tuesday',
            'الأربعاء': 'Wednesday',
            'الجمعة': 'Friday',
            'السبت': 'Saturday',
            'الأحد': 'Sunday'
        }
        
        if self.service_days_checkbox.isChecked():
            if self.monday_check.isChecked():
                service_days.append('Monday')
            if self.tuesday_check.isChecked():
                service_days.append('Tuesday')
            if self.wednesday_check.isChecked():
                service_days.append('Wednesday')
            if self.friday_check.isChecked():
                service_days.append('Friday')
            if self.saturday_check.isChecked():
                service_days.append('Saturday')
            if self.sunday_check.isChecked():
                service_days.append('Sunday')
        
        return service_days
    
    def get_service_days_only(self):
        """هل تم اختيار أيام الخدمة فقط؟"""
        return self.service_days_checkbox.isChecked()

class ComparisonReportTab(QWidget):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.excel_handler = ExcelHandler()
        self.comparison_data = {}
        self.dates_in_range = []
        self.selected_class = None
        self.service_days_only = False
        self.service_days_list = ['Thursday']  # الخميس كأساسي
        self.daily_stats_for_export = []  # تخزين الإحصائيات اليومية للتصدير
        self.setup_ui()
        
    def setup_ui(self):
        main_layout = QHBoxLayout()
        main_layout.setSpacing(16)
        main_layout.setContentsMargins(24, 24, 24, 24)
        
        # ═══════════════════════════════════════════════════════════════
        # LEFT PANEL: Controls & Stats
        # ═══════════════════════════════════════════════════════════════
        left_panel = QFrame()
        left_panel.setProperty("class", "dashboard-card")
        left_panel.setFixedWidth(350)
        # Header
        header_widget = QWidget()
        header = QHBoxLayout(header_widget)
        header_icon = QLabel("📊")
        header_icon.setProperty("class", "card-icon")
        header_title = QLabel("تقرير المقارنة")
        header_title.setProperty("class", "card-title")
        header.addWidget(header_icon)
        header.addWidget(header_title)
        header.addStretch()
        
        # Scroll Area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("background: transparent;")
        
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(14)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(20, 20, 20, 20)
        left_layout.addWidget(header_widget)
        
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setProperty("class", "card-separator")
        scroll_layout.addWidget(sep)
        
        # Period Selection
        period_label = QLabel("⏰ الفترة الزمنية")
        period_label.setProperty("class", "input-label")
        scroll_layout.addWidget(period_label)
        
        self.select_period_btn = QPushButton("📅 اختيار الفترة")
        self.select_period_btn.setProperty("class", "dashboard-combo")
        self.select_period_btn.setMinimumHeight(40)
        self.select_period_btn.clicked.connect(self.select_date_range)
        scroll_layout.addWidget(self.select_period_btn)
        
        self.period_label = QLabel("الفترة: غير محدد")
        self.period_label.setProperty("class", "stat-subtitle")
        scroll_layout.addWidget(self.period_label)
        
        self.service_days_info = QLabel("أيام الخدمة: الخميس فقط")
        self.service_days_info.setProperty("class", "filter-info-success")
        scroll_layout.addWidget(self.service_days_info)
        
        # Class Selection
        class_label = QLabel("📚 الصف للمقارنة")
        class_label.setProperty("class", "input-label")
        scroll_layout.addWidget(class_label)
        
        self.class_combo = QComboBox()
        self.class_combo.setProperty("class", "dashboard-combo")
        self.class_combo.addItems(["الكل", "الصف الأول", "الصف الثاني", "الصف الثالث"])
        self.class_combo.currentTextChanged.connect(self.on_class_changed)
        scroll_layout.addWidget(self.class_combo)
        
        # Actions
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.HLine)
        sep2.setProperty("class", "card-separator")
        scroll_layout.addWidget(sep2)
        
        actions_label = QLabel("⚡ الإجراءات")
        actions_label.setProperty("class", "input-label")
        scroll_layout.addWidget(actions_label)
        
        self.generate_btn = QPushButton("🔍 توليد التقرير")
        self.generate_btn.setProperty("class", "btn-success")
        self.generate_btn.setMinimumHeight(40)
        self.generate_btn.clicked.connect(self.generate_comparison_report)
        self.generate_btn.setEnabled(False)
        scroll_layout.addWidget(self.generate_btn)
        
        self.export_btn = QPushButton("📤 تصدير التقرير")
        self.export_btn.setProperty("class", "btn-secondary")
        self.export_btn.clicked.connect(self.export_comparison_report)
        self.export_btn.setEnabled(False)
        scroll_layout.addWidget(self.export_btn)
        
        # Stats
        sep3 = QFrame()
        sep3.setFrameShape(QFrame.HLine)
        sep3.setProperty("class", "card-separator")
        scroll_layout.addWidget(sep3)
        
        stats_label = QLabel("📈 الإحصائيات العامة")
        stats_label.setProperty("class", "input-label")
        scroll_layout.addWidget(stats_label)
        
        self.total_days_label = self.create_stat_frame("📅 عدد الأيام", "0")
        self.total_children_label = self.create_stat_frame("👥 عدد الأطفال", "0")
        self.avg_attendance_label = self.create_stat_frame("📊 متوسط الحضور", "0%")
        self.best_day_label = self.create_stat_frame("⭐ أفضل يوم", "-")
        self.worst_day_label = self.create_stat_frame("⚠️ أسوأ يوم", "-")
        
        scroll_layout.addWidget(self.total_days_label)
        scroll_layout.addWidget(self.total_children_label)
        scroll_layout.addWidget(self.avg_attendance_label)
        scroll_layout.addWidget(self.best_day_label)
        scroll_layout.addWidget(self.worst_day_label)
        
        scroll_layout.addStretch()
        
        scroll.setWidget(scroll_content)
        left_layout.addWidget(scroll)
        
        main_layout.addWidget(left_panel)
        
        # ═══════════════════════════════════════════════════════════════
        # RIGHT PANEL: Results
        # ═══════════════════════════════════════════════════════════════
        right_panel = QFrame()
        right_panel.setProperty("class", "dashboard-card")
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(12)
        right_layout.setContentsMargins(20, 20, 20, 20)
        
        # Tabs
        self.tabs = QTabWidget()
        
        self.summary_tab = QWidget()
        self.setup_summary_tab()
        
        self.details_tab = QWidget()
        self.setup_details_tab()
        
        self.comparison_tab = QWidget()
        self.setup_comparison_tab()
        
        self.followup_tab = QWidget()
        self.setup_followup_tab()
        
        self.tabs.addTab(self.summary_tab, "📋 الملخص")
        self.tabs.addTab(self.details_tab, "📊 التفاصيل")
        self.tabs.addTab(self.comparison_tab, "📈 المقارنة")
        self.tabs.addTab(self.followup_tab, "🎯 المتابعة")
        
        right_layout.addWidget(self.tabs)
        main_layout.addWidget(right_panel, 1)
        
        self.setLayout(main_layout)
    
    def create_stat_frame(self, label_text, value_text):
        """Create a stat frame widget"""
        frame = QFrame()
        frame.setStyleSheet("background: #0f0f1a; border-radius: 6px;")
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(12, 8, 12, 8)
        
        lbl = QLabel(label_text)
        lbl.setProperty("class", "stat-subtitle")
        
        val = QLabel(value_text)
        val.setStyleSheet("font-weight: 600; color: #e2e8f0;")
        
        layout.addWidget(lbl)
        layout.addStretch()
        layout.addWidget(val)
        
        return frame
    
    def update_stat_frame(self, frame, value):
        """Update stat frame value"""
        labels = frame.findChildren(QLabel)
        if len(labels) >= 2:
            labels[1].setText(str(value))
    
    def setup_summary_tab(self):
        """إعداد تبويب الملخص العام"""
        layout = QVBoxLayout()
        
        # معلومات التصفية
        self.filter_info_label = QLabel("جاري عرض جميع الأيام")
        self.filter_info_label.setProperty("class", "filter-info-danger")

        layout.addWidget(self.filter_info_label)
        
        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(8)
        self.summary_table.setHorizontalHeaderLabels([
            "التاريخ", "اليوم", "نوع اليوم", "إجمالي", "حاضر", "غائب", "النسبة", "التغيير"
        ])
        
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.summary_table.setAlternatingRowColors(True)
        self.summary_table.setSortingEnabled(True)
        
        layout.addWidget(self.summary_table)
        self.summary_tab.setLayout(layout)
    
    def setup_details_tab(self):
        """إعداد تبويب التفاصيل اليومية"""
        layout = QVBoxLayout()
        
        # معلومات التصفية
        self.details_filter_label = QLabel("جاري عرض جميع الأيام")
        self.details_filter_label.setProperty("class", "filter-info-danger")

        layout.addWidget(self.details_filter_label)
        
        # اختيار التاريخ
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("اختر تاريخ:"))
        self.details_date_combo = QComboBox()
        self.details_date_combo.currentTextChanged.connect(self.load_day_details)
        date_layout.addWidget(self.details_date_combo)
        date_layout.addStretch()
        
        # جدول التفاصيل
        self.details_table = QTableWidget()
        self.details_table.setColumnCount(10)
        self.details_table.setHorizontalHeaderLabels([
            "الكود", "الاسم", "الصف", "الحالة", "الوقت", "العنوان", "هاتف الطفل", 
            "هاتف الأب", "هاتف الأم", "ملاحظات"
        ])
        
        self.details_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.details_table.setAlternatingRowColors(True)
        self.details_table.setSortingEnabled(True)
        
        # ضبط أبعاد الأعمدة
        self.details_table.setColumnWidth(0, 80)   # الكود
        self.details_table.setColumnWidth(1, 150)  # الاسم
        self.details_table.setColumnWidth(2, 80)   # الصف
        self.details_table.setColumnWidth(3, 80)   # الحالة
        self.details_table.setColumnWidth(4, 80)   # الوقت
        self.details_table.setColumnWidth(5, 150)  # العنوان
        self.details_table.setColumnWidth(6, 100)  # هاتف الطفل
        self.details_table.setColumnWidth(7, 100)  # هاتف الأب
        self.details_table.setColumnWidth(8, 100)  # هاتف الأم
        self.details_table.setColumnWidth(9, 150)  # ملاحظات
        
        layout.addLayout(date_layout)
        layout.addWidget(self.details_table)
        self.details_tab.setLayout(layout)
    
    def setup_comparison_tab(self):
        """إعداد تبويب مقارنة الأطفال"""
        layout = QVBoxLayout()
        
        # حقل البحث
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 بحث:"))
        self.comparison_search = QLineEdit()
        self.comparison_search.setPlaceholderText("ابحث بالاسم أو الكود...")
        self.comparison_search.textChanged.connect(self.filter_comparison_table)
        search_layout.addWidget(self.comparison_search)
        search_layout.addStretch()
        
        # جدول المقارنة
        self.comparison_table = QTableWidget()
        self.comparison_table.setColumnCount(15)
        self.comparison_table.setHorizontalHeaderLabels([
            "الكود", "الاسم", "الصف", "إجمالي الأيام", "أيام الحضور", "أيام الغياب", 
            "نسبة الحضور", "مستوى الحضور", "أول حضور", "آخر حضور", "متوسط الوقت",
            "أفضل نسبة", "أسوأ نسبة", "أيام متتالية", "توصية"
        ])
        
        self.comparison_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.comparison_table.setAlternatingRowColors(True)
        self.comparison_table.setSortingEnabled(True)
        
        # ضبط أبعاد الأعمدة
        column_widths = [80, 150, 80, 90, 90, 90, 80, 100, 90, 90, 90, 90, 90, 90, 120]
        for i, width in enumerate(column_widths):
            self.comparison_table.setColumnWidth(i, width)
        
        layout.addLayout(search_layout)
        layout.addWidget(self.comparison_table)
        self.comparison_tab.setLayout(layout)
    
    def setup_followup_tab(self):
        """إعداد تبويب توصيات المتابعة"""
        layout = QVBoxLayout()
        
        # إحصائيات المتابعة
        followup_stats_layout = QHBoxLayout()
        
        self.excellent_label = QLabel("ممتاز: 0")
        self.excellent_label.setProperty("class", "status-excellent")

        
        self.good_label = QLabel("جيد: 0")
        self.good_label.setProperty("class", "status-good")

        
        self.need_followup_label = QLabel("يحتاج متابعة: 0")
        self.need_followup_label.setProperty("class", "status-warning")

        
        self.urgent_label = QLabel("متابعة عاجلة: 0")
        self.urgent_label.setProperty("class", "status-urgent")

        
        followup_stats_layout.addWidget(self.excellent_label)
        followup_stats_layout.addWidget(self.good_label)
        followup_stats_layout.addWidget(self.need_followup_label)
        followup_stats_layout.addWidget(self.urgent_label)
        followup_stats_layout.addStretch()
        
        # جدول المتابعة
        self.followup_table = QTableWidget()
        self.followup_table.setColumnCount(8)
        self.followup_table.setHorizontalHeaderLabels([
            "الكود", "الاسم", "الصف", "نسبة الحضور", "المستوى", 
            "أيام الغياب المتتالية", "آخر حضور", "توصية المتابعة"
        ])
        
        self.followup_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.followup_table.setAlternatingRowColors(True)
        self.followup_table.setSortingEnabled(True)
        
        layout.addLayout(followup_stats_layout)
        layout.addWidget(self.followup_table)
        self.followup_tab.setLayout(layout)
    
    def select_date_range(self):
        """فتح نافذة اختيار الفترة الزمنية"""
        dialog = DateRangeDialog(self)
        if dialog.exec_():
            start_date = dialog.start_date.date().toString("yyyy-MM-dd")
            end_date = dialog.end_date.date().toString("yyyy-MM-dd")
            
            # الحصول على خيار أيام الخدمة فقط
            self.service_days_only = dialog.get_service_days_only()
            self.service_days_list = dialog.get_selected_service_days()
            
            # تحديث التسمية
            period_text = f"الفترة: {start_date} إلى {end_date}"
            if self.service_days_only:
                period_text += " | أيام الخدمة فقط"
            
            self.period_label.setText(period_text)
            
            # تحديث معلومات أيام الخدمة
            service_days_text = "أيام الخدمة: "
            arabic_days = {
                'Monday': 'الإثنين',
                'Tuesday': 'الثلاثاء',
                'Wednesday': 'الأربعاء',
                'Thursday': 'الخميس',
                'Friday': 'الجمعة',
                'Saturday': 'السبت',
                'Sunday': 'الأحد'
            }
            
            service_days_text += ", ".join([arabic_days.get(day, day) for day in self.service_days_list])
            self.service_days_info.setText(service_days_text)
            
            self.generate_btn.setEnabled(True)
            
            # حفظ التواريخ المحددة
            self.start_date = start_date
            self.end_date = end_date
    
    def on_class_changed(self, class_name):
        """عند تغيير الصف المحدد"""
        self.selected_class = class_name if class_name != "الكل" else None
    
    def generate_comparison_report(self):
        """توليد تقرير المقارنة"""
        try:
            if not hasattr(self, 'start_date'):
                QMessageBox.warning(self, "تحذير", "يرجى اختيار الفترة الزمنية أولاً")
                return
            
            # عرض تقدم العمل
            progress = QMessageBox(self)
            progress.setWindowTitle("جاري المعالجة")
            progress.setText("جاري توليد تقرير المقارنة...")
            progress.show()
            
            # الحصول على جميع التواريخ في الفترة
            start = datetime.strptime(self.start_date, "%Y-%m-%d")
            end = datetime.strptime(self.end_date, "%Y-%m-%d")
            
            all_dates_in_range = []
            current = start
            while current <= end:
                all_dates_in_range.append(current.strftime("%Y-%m-%d"))
                current += timedelta(days=1)
            
            # فلترة التواريخ إذا تم اختيار أيام الخدمة فقط
            if self.service_days_only:
                filtered_dates = []
                for date in all_dates_in_range:
                    date_obj = datetime.strptime(date, "%Y-%m-%d")
                    day_english = date_obj.strftime("%A")  # Monday, Tuesday, etc.
                    
                    # التحقق إذا كان اليوم في قائمة أيام الخدمة
                    if day_english in self.service_days_list:
                        filtered_dates.append(date)
                
                self.dates_in_range = filtered_dates
                
                if not self.dates_in_range:
                    progress.close()
                    QMessageBox.warning(self, "تحذير", "لا توجد أيام خدمة في الفترة المحددة")
                    return
            else:
                self.dates_in_range = all_dates_in_range
            
            # تحديث معلومات التصفية في الواجهة
            self.update_filter_labels()
            
            # الحصول على الأطفال حسب الصف المحدد
            if self.selected_class:
                children = self.db.get_children_by_class(self.selected_class)
            else:
                children = self.db.get_all_children()
            
            # تهيئة بيانات المقارنة
            self.comparison_data = {}
            daily_stats = []
            
            for child in children:
                child_code = str(child.get('code', '')).strip()
                if not child_code:
                    continue
                
                # تهيئة بيانات الطفل
                self.comparison_data[child_code] = {
                    'info': child,
                    'attendance': {},
                    'stats': {
                        'total_days': 0,
                        'present_days': 0,
                        'absent_days': 0,
                        'attendance_rate': 0,
                        'first_attendance': None,
                        'last_attendance': None,
                        'attendance_times': [],
                        'consecutive_absent': 0,
                        'max_consecutive_absent': 0,
                        'best_rate': 0,
                        'worst_rate': 100
                    }
                }
            
            # تحليل بيانات كل يوم
            for date in self.dates_in_range:
                # الحصول على بيانات الحضور لهذا اليوم
                data = self.db.load_data()
                attendance_data = data.get('attendance', {}).get(date, {})
                
                # إحصائيات اليوم
                present_codes = set(attendance_data.keys())
                total_children = len(children)
                present_count = len(present_codes)
                absent_count = total_children - present_count
                attendance_rate = (present_count / total_children * 100) if total_children > 0 else 0
                
                # معلومات اليوم
                date_obj = datetime.strptime(date, "%Y-%m-%d")
                day_name = self.db.arabic_days.get(date_obj.strftime("%A"), date_obj.strftime("%A"))
                
                # التحقق إذا كان يوم خدمة
                day_english = date_obj.strftime("%A")
                is_service_day = (day_english in self.service_days_list)
                
                daily_stats.append({
                    'date': date,
                    'day_name': day_name,
                    'is_service_day': is_service_day,
                    'total': total_children,
                    'present': present_count,
                    'absent': absent_count,
                    'rate': attendance_rate
                })
                
                # تحديث إحصائيات كل طفل
                for child_code, child_data in self.comparison_data.items():
                    is_present = child_code in present_codes
                    arrival_time = attendance_data.get(child_code, None)
                    
                    child_data['attendance'][date] = {
                        'present': is_present,
                        'time': arrival_time
                    }
                    
                    if is_present:
                        child_data['stats']['present_days'] += 1
                        if arrival_time:
                            child_data['stats']['attendance_times'].append(arrival_time)
                        
                        # تحديث أول وآخر حضور
                        if not child_data['stats']['first_attendance']:
                            child_data['stats']['first_attendance'] = date
                        child_data['stats']['last_attendance'] = date
                        
                        # إعادة تعيين الغياب المتتالي
                        child_data['stats']['consecutive_absent'] = 0
                    else:
                        child_data['stats']['absent_days'] += 1
                        child_data['stats']['consecutive_absent'] += 1
                        child_data['stats']['max_consecutive_absent'] = max(
                            child_data['stats']['max_consecutive_absent'],
                            child_data['stats']['consecutive_absent']
                        )
            
            # حفظ الإحصائيات اليومية للتصدير
            self.daily_stats_for_export = daily_stats
            
            # حساب الإحصائيات النهائية لكل طفل
            for child_code, child_data in self.comparison_data.items():
                stats = child_data['stats']
                stats['total_days'] = len(self.dates_in_range)
                
                if stats['total_days'] > 0:
                    stats['attendance_rate'] = (stats['present_days'] / stats['total_days']) * 100
                
                # حساب متوسط وقت الحضور
                if stats['attendance_times']:
                    total_minutes = 0
                    for time_str in stats['attendance_times']:
                        try:
                            hours, minutes = map(int, time_str.split(':'))
                            total_minutes += hours * 60 + minutes
                        except:
                            pass
                    
                    avg_minutes = total_minutes // len(stats['attendance_times'])
                    stats['avg_time'] = f"{avg_minutes // 60:02d}:{avg_minutes % 60:02d}"
                else:
                    stats['avg_time'] = "00:00"
            
            # تحديث الجداول
            self.update_summary_table(daily_stats)
            self.update_comparison_table()
            self.update_followup_table()
            self.update_details_combo()
            
            # تحديث الإحصائيات العامة
            self.update_general_stats(daily_stats)
            
            # تفعيل زر التصدير
            self.export_btn.setEnabled(True)
            
            progress.close()
            
            # رسالة النجاح مع معلومات إضافية
            message = f"تم توليد تقرير المقارنة بنجاح!\n"
            message += f"الفترة: {self.start_date} إلى {self.end_date}\n"
            message += f"عدد الأيام: {len(self.dates_in_range)}\n"
            message += f"عدد الأطفال: {len(self.comparison_data)}\n"
            
            if self.service_days_only:
                message += f"تم عرض أيام الخدمة فقط\n"
                message += f"أيام الخدمة: {', '.join(self.service_days_list)}"
            
            QMessageBox.information(self, "تم التوليد", message)
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء توليد التقرير: {str(e)}")
    
    def update_filter_labels(self):
        """تحديث تسميات التصفية في الواجهة"""
        if self.service_days_only:
            filter_text = f"📅 جاري عرض أيام الخدمة فقط ({len(self.dates_in_range)} يوم)"
            arabic_days = {
                'Monday': 'الإثنين',
                'Tuesday': 'الثلاثاء',
                'Wednesday': 'الأربعاء',
                'Thursday': 'الخميس',
                'Friday': 'الجمعة',
                'Saturday': 'السبت',
                'Sunday': 'الأحد'
            }
            days_text = ", ".join([arabic_days.get(day, day) for day in self.service_days_list])
            filter_text += f"\n🗓️ الأيام المعروضة: {days_text}"
            
            self.filter_info_label.setProperty("class", "filter-info-success")
            self.filter_info_label.style().unpolish(self.filter_info_label)
            self.filter_info_label.style().polish(self.filter_info_label)

        else:
            filter_text = "📅 جاري عرض جميع الأيام في الفترة"
            self.filter_info_label.setProperty("class", "filter-info-primary")
            self.filter_info_label.style().unpolish(self.filter_info_label)
            self.filter_info_label.style().polish(self.filter_info_label)

        
        self.filter_info_label.setText(filter_text)
        self.details_filter_label.setText(filter_text)
    
    def update_summary_table(self, daily_stats):
        """تحديث جدول الملخص العام"""
        self.summary_table.setSortingEnabled(False)
        self.summary_table.setRowCount(0)
        
        previous_rate = None
        for i, day in enumerate(daily_stats):
            row = self.summary_table.rowCount()
            self.summary_table.insertRow(row)
            
            # التاريخ
            self.summary_table.setItem(row, 0, QTableWidgetItem(day['date']))
            
            # اليوم
            day_item = QTableWidgetItem(day['day_name'])
            if day['is_service_day']:
                day_item.setBackground(Qt.yellow)
                day_item.setToolTip("يوم الخدمة")
            self.summary_table.setItem(row, 1, day_item)
            
            # نوع اليوم
            type_item = QTableWidgetItem("يوم خدمة" if day['is_service_day'] else "يوم عادي")
            self.summary_table.setItem(row, 2, type_item)
            
            # الإجمالي
            self.summary_table.setItem(row, 3, QTableWidgetItem(str(day['total'])))
            
            # الحاضر
            present_item = QTableWidgetItem(str(day['present']))
            self.summary_table.setItem(row, 4, present_item)
            
            # الغائب
            absent_item = QTableWidgetItem(str(day['absent']))
            self.summary_table.setItem(row, 5, absent_item)
            
            # النسبة
            rate_item = QTableWidgetItem(f"{day['rate']:.1f}%")
            
            # تلوين حسب النسبة
            if day['rate'] >= 80:
                rate_item.setBackground(Qt.green)
                rate_item.setForeground(Qt.white)
            elif day['rate'] >= 60:
                rate_item.setBackground(Qt.yellow)
            elif day['rate'] >= 40:
                rate_item.setBackground(QColor(255, 165, 0))  # برتقالي
            else:
                rate_item.setBackground(Qt.red)
                rate_item.setForeground(Qt.white)
            
            self.summary_table.setItem(row, 6, rate_item)
            
            # التغيير عن اليوم السابق
            if previous_rate is not None:
                change = day['rate'] - previous_rate
                change_item = QTableWidgetItem(f"{change:+.1f}%")
                
                if change > 0:
                    change_item.setBackground(Qt.green)
                    change_item.setForeground(Qt.white)
                elif change < 0:
                    change_item.setBackground(Qt.red)
                    change_item.setForeground(Qt.white)
                
                self.summary_table.setItem(row, 7, change_item)
            else:
                self.summary_table.setItem(row, 7, QTableWidgetItem("-"))
            
            previous_rate = day['rate']
        
        self.summary_table.setSortingEnabled(True)
    
    def update_comparison_table(self):
        """تحديث جدول مقارنة الأطفال"""
        self.comparison_table.setSortingEnabled(False)
        self.comparison_table.setRowCount(0)
        
        for child_code, child_data in self.comparison_data.items():
            child_info = child_data['info']
            stats = child_data['stats']
            
            row = self.comparison_table.rowCount()
            self.comparison_table.insertRow(row)
            
            # الكود
            self.comparison_table.setItem(row, 0, QTableWidgetItem(child_code))
            
            # الاسم
            self.comparison_table.setItem(row, 1, QTableWidgetItem(child_info.get('name', 'غير محدد')))
            
            # الصف
            self.comparison_table.setItem(row, 2, QTableWidgetItem(child_info.get('class', 'غير محدد')))
            
            # إجمالي الأيام
            self.comparison_table.setItem(row, 3, QTableWidgetItem(str(stats['total_days'])))
            
            # أيام الحضور
            present_item = QTableWidgetItem(str(stats['present_days']))
            self.comparison_table.setItem(row, 4, present_item)
            
            # أيام الغياب
            absent_item = QTableWidgetItem(str(stats['absent_days']))
            self.comparison_table.setItem(row, 5, absent_item)
            
            # نسبة الحضور
            rate_item = QTableWidgetItem(f"{stats['attendance_rate']:.1f}%")
            
            # تلوين حسب النسبة
            if stats['attendance_rate'] >= 80:
                rate_item.setBackground(Qt.green)
                rate_item.setForeground(Qt.white)
            elif stats['attendance_rate'] >= 60:
                rate_item.setBackground(Qt.yellow)
            elif stats['attendance_rate'] >= 40:
                rate_item.setBackground(QColor(255, 165, 0))
            else:
                rate_item.setBackground(Qt.red)
                rate_item.setForeground(Qt.white)
            
            self.comparison_table.setItem(row, 6, rate_item)
            
            # مستوى الحضور
            level = self.get_attendance_level(stats['attendance_rate'])
            level_item = QTableWidgetItem(level)
            self.comparison_table.setItem(row, 7, level_item)
            
            # أول حضور
            self.comparison_table.setItem(row, 8, QTableWidgetItem(stats.get('first_attendance', '-')))
            
            # آخر حضور
            self.comparison_table.setItem(row, 9, QTableWidgetItem(stats.get('last_attendance', '-')))
            
            # متوسط الوقت
            self.comparison_table.setItem(row, 10, QTableWidgetItem(stats.get('avg_time', '00:00')))
            
            # أفضل نسبة (سيتم حسابها)
            best_rate = self.calculate_best_attendance_rate(child_data['attendance'])
            self.comparison_table.setItem(row, 11, QTableWidgetItem(f"{best_rate:.1f}%"))
            
            # أسوأ نسبة
            worst_rate = self.calculate_worst_attendance_rate(child_data['attendance'])
            self.comparison_table.setItem(row, 12, QTableWidgetItem(f"{worst_rate:.1f}%"))
            
            # أيام غياب متتالية
            self.comparison_table.setItem(row, 13, QTableWidgetItem(str(stats['max_consecutive_absent'])))
            
            # توصية
            recommendation = self.get_recommendation(stats['attendance_rate'], stats['max_consecutive_absent'])
            rec_item = QTableWidgetItem(recommendation)
            
            # تلوين التوصية
            if "ممتاز" in recommendation:
                rec_item.setBackground(Qt.green)
                rec_item.setForeground(Qt.white)
            elif "جيد" in recommendation:
                rec_item.setBackground(Qt.blue)
                rec_item.setForeground(Qt.white)
            elif "يحتاج متابعة" in recommendation:
                rec_item.setBackground(Qt.yellow)
            else:
                rec_item.setBackground(Qt.red)
                rec_item.setForeground(Qt.white)
            
            self.comparison_table.setItem(row, 14, rec_item)
        
        self.comparison_table.setSortingEnabled(True)
    
    def update_followup_table(self):
        """تحديث جدول توصيات المتابعة"""
        self.followup_table.setSortingEnabled(False)
        self.followup_table.setRowCount(0)
        
        # عدادات المستويات
        level_counts = {'ممتاز': 0, 'جيد': 0, 'يحتاج متابعة': 0, 'متابعة عاجلة': 0}
        
        for child_code, child_data in self.comparison_data.items():
            child_info = child_data['info']
            stats = child_data['stats']
            
            row = self.followup_table.rowCount()
            self.followup_table.insertRow(row)
            
            # الكود
            self.followup_table.setItem(row, 0, QTableWidgetItem(child_code))
            
            # الاسم
            self.followup_table.setItem(row, 1, QTableWidgetItem(child_info.get('name', 'غير محدد')))
            
            # الصف
            self.followup_table.setItem(row, 2, QTableWidgetItem(child_info.get('class', 'غير محدد')))
            
            # نسبة الحضور
            rate_item = QTableWidgetItem(f"{stats['attendance_rate']:.1f}%")
            self.followup_table.setItem(row, 3, rate_item)
            
            # المستوى
            level = self.get_attendance_level(stats['attendance_rate'])
            level_item = QTableWidgetItem(level)
            
            # تلوين حسب المستوى
            if level == "ممتاز":
                level_item.setBackground(Qt.green)
                level_item.setForeground(Qt.white)
                level_counts['ممتاز'] += 1
            elif level == "جيد":
                level_item.setBackground(Qt.blue)
                level_item.setForeground(Qt.white)
                level_counts['جيد'] += 1
            elif level == "يحتاج متابعة":
                level_item.setBackground(Qt.yellow)
                level_counts['يحتاج متابعة'] += 1
            else:
                level_item.setBackground(Qt.red)
                level_item.setForeground(Qt.white)
                level_counts['متابعة عاجلة'] += 1
            
            self.followup_table.setItem(row, 4, level_item)
            
            # أيام الغياب المتتالية
            self.followup_table.setItem(row, 5, QTableWidgetItem(str(stats['max_consecutive_absent'])))
            
            # آخر حضور
            self.followup_table.setItem(row, 6, QTableWidgetItem(stats.get('last_attendance', '-')))
            
            # توصية المتابعة
            recommendation = self.get_followup_recommendation(
                stats['attendance_rate'], 
                stats['max_consecutive_absent'],
                stats.get('last_attendance', None)
            )
            self.followup_table.setItem(row, 7, QTableWidgetItem(recommendation))
        
        self.followup_table.setSortingEnabled(True)
        
        # تحديث عدادات المستويات
        self.excellent_label.setText(f"ممتاز: {level_counts['ممتاز']}")
        self.good_label.setText(f"جيد: {level_counts['جيد']}")
        self.need_followup_label.setText(f"يحتاج متابعة: {level_counts['يحتاج متابعة']}")
        self.urgent_label.setText(f"متابعة عاجلة: {level_counts['متابعة عاجلة']}")
    
    def update_details_combo(self):
        """تحديث قائمة التواريخ في تبويب التفاصيل"""
        self.details_date_combo.clear()
        for date in self.dates_in_range:
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            day_name = self.db.arabic_days.get(date_obj.strftime("%A"), date_obj.strftime("%A"))
            self.details_date_combo.addItem(f"{date} ({day_name})", date)
    
    def load_day_details(self):
        """تحميل تفاصيل يوم محدد"""
        try:
            if self.details_date_combo.currentIndex() < 0:
                return
            
            selected_date = self.details_date_combo.currentData()
            if not selected_date or not self.comparison_data:
                return
            
            self.details_table.setSortingEnabled(False)
            self.details_table.setRowCount(0)
            
            for child_code, child_data in self.comparison_data.items():
                child_info = child_data['info']
                attendance_info = child_data['attendance'].get(selected_date, {})
                is_present = attendance_info.get('present', False)
                arrival_time = attendance_info.get('time', 'غير محدد')
                
                row = self.details_table.rowCount()
                self.details_table.insertRow(row)
                
                # الكود
                self.details_table.setItem(row, 0, QTableWidgetItem(child_code))
                
                # الاسم
                self.details_table.setItem(row, 1, QTableWidgetItem(child_info.get('name', 'غير محدد')))
                
                # الصف
                self.details_table.setItem(row, 2, QTableWidgetItem(child_info.get('class', 'غير محدد')))
                
                # الحالة
                status_item = QTableWidgetItem("حاضر" if is_present else "غائب")
                if is_present:
                    status_item.setBackground(Qt.green)
                    status_item.setForeground(Qt.white)
                else:
                    status_item.setBackground(Qt.red)
                    status_item.setForeground(Qt.white)
                self.details_table.setItem(row, 3, status_item)
                
                # الوقت
                time_item = QTableWidgetItem(arrival_time if is_present else "-")
                self.details_table.setItem(row, 4, time_item)
                
                # العنوان
                address = f"{child_info.get('عماره', '')} - {child_info.get('شارع', '')}"
                self.details_table.setItem(row, 5, QTableWidgetItem(address if address.strip() != '-' else 'غير محدد'))
                
                # هواتف
                self.details_table.setItem(row, 6, QTableWidgetItem(child_info.get('موبيل الولد', 'غير محدد')))
                self.details_table.setItem(row, 7, QTableWidgetItem(child_info.get('موبايل الاب', 'غير محدد')))
                self.details_table.setItem(row, 8, QTableWidgetItem(child_info.get('موبايل الام', 'غير محدد')))
                
                # ملاحظات
                notes = child_info.get('ملاحظات', 'غير محدد')
                if not is_present and child_data['stats']['max_consecutive_absent'] > 2:
                    notes += " (غياب متكرر)"
                self.details_table.setItem(row, 9, QTableWidgetItem(notes))
            
            self.details_table.setSortingEnabled(True)
            
        except Exception as e:
            print(f"Error loading day details: {str(e)}")
    
    def filter_comparison_table(self, search_text):
        """تصفية جدول المقارنة حسب البحث"""
        search_text = search_text.lower().strip()
        
        for row in range(self.comparison_table.rowCount()):
            show_row = False
            
            # البحث في الكود والاسم
            code = self.comparison_table.item(row, 0).text().lower()
            name = self.comparison_table.item(row, 1).text().lower()
            
            if search_text in code or search_text in name:
                show_row = True
            
            # إظهار أو إخفاء الصف
            self.comparison_table.setRowHidden(row, not show_row)
    
    def update_general_stats(self, daily_stats):
        """تحديث الإحصائيات العامة"""
        if not daily_stats:
            return
        
        # إحصائيات عامة
        total_days = len(daily_stats)
        total_children = len(self.comparison_data)
        avg_attendance = sum(day['rate'] for day in daily_stats) / total_days if total_days > 0 else 0
        
        # أفضل وأسوأ يوم
        best_day = max(daily_stats, key=lambda x: x['rate'])
        worst_day = min(daily_stats, key=lambda x: x['rate'])
        
        self.update_stat_frame(self.total_days_label, str(total_days))
        self.update_stat_frame(self.total_children_label, str(total_children))
        self.update_stat_frame(self.avg_attendance_label, f"{avg_attendance:.1f}%")
        self.update_stat_frame(self.best_day_label, f"{best_day['date']} ({best_day['rate']:.1f}%)")
        self.update_stat_frame(self.worst_day_label, f"{worst_day['date']} ({worst_day['rate']:.1f}%)")
    
    def get_attendance_level(self, rate):
        """الحصول على مستوى الحضور"""
        if rate >= 90:
            return "ممتاز"
        elif rate >= 75:
            return "جيد جداً"
        elif rate >= 60:
            return "جيد"
        elif rate >= 40:
            return "مقبول"
        else:
            return "ضعيف"
    
    def calculate_best_attendance_rate(self, attendance_data):
        """حساب أفضل نسبة حضور للطفل"""
        if not attendance_data:
            return 0
        
        # حساب النسبة لكل أسبوع
        dates = sorted(attendance_data.keys())
        best_rate = 0
        
        for i in range(0, len(dates), 7):
            week_dates = dates[i:i+7]
            present_count = sum(1 for date in week_dates if attendance_data[date]['present'])
            
            if week_dates:
                week_rate = (present_count / len(week_dates)) * 100
                best_rate = max(best_rate, week_rate)
        
        return best_rate
    
    def calculate_worst_attendance_rate(self, attendance_data):
        """حساب أسوأ نسبة حضور للطفل"""
        if not attendance_data:
            return 100
        
        dates = sorted(attendance_data.keys())
        worst_rate = 100
        
        for i in range(0, len(dates), 7):
            week_dates = dates[i:i+7]
            present_count = sum(1 for date in week_dates if attendance_data[date]['present'])
            
            if week_dates:
                week_rate = (present_count / len(week_dates)) * 100
                worst_rate = min(worst_rate, week_rate)
        
        return worst_rate
    
    def get_recommendation(self, rate, consecutive_absent):
        """الحصول على توصية"""
        if rate >= 80:
            return "ممتاز - لا يحتاج متابعة"
        elif rate >= 60:
            return "جيد - متابعة خفيفة"
        elif consecutive_absent > 3:
            return "يحتاج متابعة فورية"
        elif rate >= 40:
            return "يحتاج متابعة منتظمة"
        else:
            return "متابعة مكثفة وعاجلة"
    
    def get_followup_recommendation(self, rate, consecutive_absent, last_attendance):
        """الحصول على توصية متابعة مفصلة"""
        recommendations = []
        
        if rate < 50:
            recommendations.append("زيارة منزلية")
        
        if consecutive_absent > 2:
            recommendations.append("اتصال هاتفي مع الأهل")
        
        if last_attendance:
            last_date = datetime.strptime(last_attendance, "%Y-%m-%d")
            days_since_last = (datetime.now() - last_date).days
            
            if days_since_last > 14:
                recommendations.append("متابعة شخصية مع الطفل")
        
        if rate < 30:
            recommendations.append("اجتماع مع ولي الأمر")
        
        if not recommendations:
            return "لا يحتاج متابعة خاصة"
        
        return " + ".join(recommendations)
    
    def export_comparison_report(self):
        """تصدير تقرير المقارنة"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            if not self.comparison_data:
                QMessageBox.warning(self, "تحذير", "لا توجد بيانات للتصدير")
                return
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ تقرير المقارنة", 
                f"تقرير_مقارنة_الحضور_{self.start_date}_إلى_{self.end_date}.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                # إنشاء ملف Excel
                wb = Workbook()
                
                # ورقة الملخص
                ws_summary = wb.active
                ws_summary.title = "الملخص العام"
                
                # إضافة العناوين
                titles = [
                    f"تقرير مقارنة الحضور - الفترة من {self.start_date} إلى {self.end_date}",
                    f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    f"عدد الأيام: {len(self.dates_in_range)}",
                    f"عدد الأطفال: {len(self.comparison_data)}"
                ]
                
                if self.service_days_only:
                    titles.append(f"أيام الخدمة: {', '.join(self.service_days_list)}")
                
                for i, title in enumerate(titles, 1):
                    ws_summary[f'A{i}'] = title
                    ws_summary[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                    ws_summary[f'A{i}'].alignment = Alignment(horizontal='center')
                
                # دمج الخلايا
                ws_summary.merge_cells('A1:H1')
                ws_summary.merge_cells('A2:H2')
                ws_summary.merge_cells('A3:H3')
                ws_summary.merge_cells('A4:H4')
                if self.service_days_only:
                    ws_summary.merge_cells('A5:H5')
                
                # إضافة رؤوس أعمدة الملخص
                summary_headers = ['التاريخ', 'اليوم', 'نوع اليوم', 'إجمالي الأطفال', 'الحضور', 'الغياب', 'النسبة', 'الحالة']
                
                header_row = 6 if not self.service_days_only else 7
                for col, header in enumerate(summary_headers, 1):
                    cell = ws_summary.cell(row=header_row, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                
                # إضافة بيانات الملخص
                if hasattr(self, 'daily_stats_for_export'):
                    data_row = header_row + 1
                    for day in self.daily_stats_for_export:
                        ws_summary.cell(row=data_row, column=1, value=day['date'])
                        ws_summary.cell(row=data_row, column=2, value=day['day_name'])
                        ws_summary.cell(row=data_row, column=3, value="يوم خدمة" if day['is_service_day'] else "يوم عادي")
                        ws_summary.cell(row=data_row, column=4, value=day['total'])
                        ws_summary.cell(row=data_row, column=5, value=day['present'])
                        ws_summary.cell(row=data_row, column=6, value=day['absent'])
                        ws_summary.cell(row=data_row, column=7, value=f"{day['rate']:.1f}%")
                        
                        data_row += 1
                
                # ورقة مقارنة الأطفال
                ws_comparison = wb.create_sheet(title="مقارنة الأطفال")
                
                # إضافة العناوين
                title_row = 1
                for i, title in enumerate(titles[:2], 1):
                    ws_comparison.cell(row=title_row, column=1, value=title)
                    ws_comparison.cell(row=title_row, column=1).font = Font(bold=True, size=14 if i == 1 else 12)
                    ws_comparison.cell(row=title_row, column=1).alignment = Alignment(horizontal='center')
                    title_row += 1
                
                if self.service_days_only:
                    ws_comparison.cell(row=title_row, column=1, value=f"أيام الخدمة: {', '.join(self.service_days_list)}")
                    ws_comparison.cell(row=title_row, column=1).font = Font(bold=True)
                    title_row += 1
                
                # دمج الخلايا للعناوين
                ws_comparison.merge_cells('A1:O1')
                ws_comparison.merge_cells('A2:O2')
                if self.service_days_only:
                    ws_comparison.merge_cells('A3:O3')
                
                # إضافة رؤوس أعمدة المقارنة
                comparison_headers = [
                    'الكود', 'الاسم', 'الصف', 'إجمالي الأيام', 'أيام الحضور', 'أيام الغياب', 
                    'نسبة الحضور', 'المستوى', 'أول حضور', 'آخر حضور', 'متوسط الوقت',
                    'أفضل نسبة', 'أسوأ نسبة', 'أيام غياب متتالية', 'التوصية'
                ]
                
                header_row = title_row + 1
                for col, header in enumerate(comparison_headers, 1):
                    cell = ws_comparison.cell(row=header_row, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                
                # إضافة بيانات المقارنة
                data_row = header_row + 1
                for child_code, child_data in self.comparison_data.items():
                    child_info = child_data['info']
                    stats = child_data['stats']
                    
                    ws_comparison.cell(row=data_row, column=1, value=child_code)
                    ws_comparison.cell(row=data_row, column=2, value=child_info.get('name', 'غير محدد'))
                    ws_comparison.cell(row=data_row, column=3, value=child_info.get('class', 'غير محدد'))
                    ws_comparison.cell(row=data_row, column=4, value=stats['total_days'])
                    ws_comparison.cell(row=data_row, column=5, value=stats['present_days'])
                    ws_comparison.cell(row=data_row, column=6, value=stats['absent_days'])
                    ws_comparison.cell(row=data_row, column=7, value=f"{stats['attendance_rate']:.1f}%")
                    ws_comparison.cell(row=data_row, column=8, value=self.get_attendance_level(stats['attendance_rate']))
                    ws_comparison.cell(row=data_row, column=9, value=stats.get('first_attendance', '-'))
                    ws_comparison.cell(row=data_row, column=10, value=stats.get('last_attendance', '-'))
                    ws_comparison.cell(row=data_row, column=11, value=stats.get('avg_time', '00:00'))
                    ws_comparison.cell(row=data_row, column=12, value=f"{self.calculate_best_attendance_rate(child_data['attendance']):.1f}%")
                    ws_comparison.cell(row=data_row, column=13, value=f"{self.calculate_worst_attendance_rate(child_data['attendance']):.1f}%")
                    ws_comparison.cell(row=data_row, column=14, value=stats['max_consecutive_absent'])
                    ws_comparison.cell(row=data_row, column=15, value=self.get_recommendation(stats['attendance_rate'], stats['max_consecutive_absent']))
                    
                    data_row += 1
                
                # ورقة المتابعة
                ws_followup = wb.create_sheet(title="توصيات المتابعة")
                
                # إضافة العناوين
                title_row = 1
                for i, title in enumerate(titles[:2], 1):
                    ws_followup.cell(row=title_row, column=1, value=title)
                    ws_followup.cell(row=title_row, column=1).font = Font(bold=True, size=14 if i == 1 else 12)
                    ws_followup.cell(row=title_row, column=1).alignment = Alignment(horizontal='center')
                    title_row += 1
                
                if self.service_days_only:
                    ws_followup.cell(row=title_row, column=1, value=f"أيام الخدمة: {', '.join(self.service_days_list)}")
                    ws_followup.cell(row=title_row, column=1).font = Font(bold=True)
                    title_row += 1
                
                ws_followup.merge_cells('A1:H1')
                ws_followup.merge_cells('A2:H2')
                if self.service_days_only:
                    ws_followup.merge_cells('A3:H3')
                
                # إحصائيات المتابعة
                stats_row = title_row + 1
                ws_followup.cell(row=stats_row, column=1, value="إحصائيات المتابعة:")
                ws_followup.cell(row=stats_row, column=1).font = Font(bold=True)
                
                # حساب الإحصائيات
                level_counts = {'ممتاز': 0, 'جيد': 0, 'يحتاج متابعة': 0, 'متابعة عاجلة': 0}
                for child_data in self.comparison_data.values():
                    level = self.get_attendance_level(child_data['stats']['attendance_rate'])
                    if level in level_counts:
                        level_counts[level] += 1
                
                stats_row += 1
                ws_followup.cell(row=stats_row, column=1, value=f"ممتاز: {level_counts['ممتاز']}")
                ws_followup.cell(row=stats_row, column=2, value=f"جيد: {level_counts['جيد']}")
                stats_row += 1
                ws_followup.cell(row=stats_row, column=1, value=f"يحتاج متابعة: {level_counts['يحتاج متابعة']}")
                ws_followup.cell(row=stats_row, column=2, value=f"متابعة عاجلة: {level_counts['متابعة عاجلة']}")
                
                # رؤوس أعمدة المتابعة
                followup_headers = ['الكود', 'الاسم', 'الصف', 'نسبة الحضور', 'المستوى', 
                                   'أيام الغياب المتتالية', 'آخر حضور', 'توصية المتابعة']
                
                header_row = stats_row + 2
                for col, header in enumerate(followup_headers, 1):
                    cell = ws_followup.cell(row=header_row, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                
                # إضافة بيانات المتابعة
                data_row = header_row + 1
                for child_code, child_data in self.comparison_data.items():
                    child_info = child_data['info']
                    stats = child_data['stats']
                    
                    ws_followup.cell(row=data_row, column=1, value=child_code)
                    ws_followup.cell(row=data_row, column=2, value=child_info.get('name', 'غير محدد'))
                    ws_followup.cell(row=data_row, column=3, value=child_info.get('class', 'غير محدد'))
                    ws_followup.cell(row=data_row, column=4, value=f"{stats['attendance_rate']:.1f}%")
                    ws_followup.cell(row=data_row, column=5, value=self.get_attendance_level(stats['attendance_rate']))
                    ws_followup.cell(row=data_row, column=6, value=stats['max_consecutive_absent'])
                    ws_followup.cell(row=data_row, column=7, value=stats.get('last_attendance', '-'))
                    ws_followup.cell(row=data_row, column=8, value=self.get_followup_recommendation(
                        stats['attendance_rate'], 
                        stats['max_consecutive_absent'],
                        stats.get('last_attendance', None)
                    ))
                    
                    data_row += 1
                
                # ضبط أبعاد الأعمدة
                for ws in [ws_summary, ws_comparison, ws_followup]:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # حفظ الملف
                wb.save(file_path)
                
                message = f"تم تصدير تقرير المقارنة بنجاح!\n\n"
                message += f"الملف: {file_path}\n"
                message += f"عدد الأطفال: {len(self.comparison_data)}\n"
                message += f"عدد الأيام: {len(self.dates_in_range)}\n"
                
                if self.service_days_only:
                    message += f"تم تصدير أيام الخدمة فقط\n"
                    message += f"أيام الخدمة: {', '.join(self.service_days_list)}"
                
                QMessageBox.information(self, "تم التصدير", message)
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء التصدير: {str(e)}")
    
    def refresh_data(self):
        """تحديث البيانات"""
        if hasattr(self, 'start_date') and hasattr(self, 'end_date'):
            self.generate_comparison_report()
        else:
            QMessageBox.information(self, "معلومة", "يرجى اختيار الفترة الزمنية أولاً")