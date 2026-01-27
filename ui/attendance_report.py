from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QTableWidget, QTableWidgetItem, QHeaderView,
                             QComboBox, QGroupBox, QPushButton, QMessageBox,
                             QSplitter, QProgressBar, QCheckBox, QFrame, QScrollArea)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon
from utils.database import DatabaseManager
from utils.excel_handler import ExcelHandler
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import arabic_reshaper
from bidi.algorithm import get_display
import os

class AttendanceReportTab(QWidget):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.excel_handler = ExcelHandler()
        self.current_attendance_data = []
        self.current_date = None
        self.setup_ui()
        self.load_dates()
        
    def setup_ui(self):
        main_layout = QHBoxLayout()
        main_layout.setSpacing(16)
        main_layout.setContentsMargins(24, 24, 24, 24)
        
        # ═══════════════════════════════════════════════════════════════
        # LEFT PANEL: Controls & Stats
        # ═══════════════════════════════════════════════════════════════
        left_panel = QFrame()
        left_panel.setProperty("class", "dashboard-card")
        left_panel.setFixedWidth(350)
        # Header
        header_widget = QWidget()
        header = QHBoxLayout(header_widget)
        header_icon = QLabel("𓃭")
        header_icon.setProperty("class", "card-icon")
        header_title = QLabel("تقرير الحضور")
        header_title.setProperty("class", "card-title")
        header.addWidget(header_icon)
        header.addWidget(header_title)
        header.addStretch()
        
        # Scroll Area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.viewport().setStyleSheet("background-color: #141420;")
        scroll.setStyleSheet("background-color: #141420; border: none;")
        
        scroll_content = QWidget()
        scroll_content.setStyleSheet("background-color: #141420;")
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(14)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(20, 20, 20, 20)
        left_layout.addWidget(header_widget)
        
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setProperty("class", "card-separator")
        scroll_layout.addWidget(sep)
        
        # Filters
        filter_label = QLabel("📋 إعدادات التقرير")
        filter_label.setProperty("class", "input-label")
        scroll_layout.addWidget(filter_label)
        
        self.date_selector = QComboBox()
        self.date_selector.setProperty("class", "dashboard-combo")
        self.date_selector.setMinimumHeight(40)
        self.date_selector.currentTextChanged.connect(self.load_attendance_data)
        scroll_layout.addWidget(self.date_selector)
        
        self.service_days_only = QCheckBox("عرض أيام الخدمة فقط")
        self.service_days_only.stateChanged.connect(self.load_dates)
        scroll_layout.addWidget(self.service_days_only)
        
        self.date_info_label = QLabel("")
        self.date_info_label.setProperty("class", "stat-subtitle")
        scroll_layout.addWidget(self.date_info_label)
        
        self.refresh_btn = QPushButton("🔄 تحديث")
        self.refresh_btn.setProperty("class", "btn-secondary")
        self.refresh_btn.clicked.connect(self.refresh_data)
        scroll_layout.addWidget(self.refresh_btn)
        
        # Stats
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.HLine)
        sep2.setProperty("class", "card-separator")
        scroll_layout.addWidget(sep2)
        
        stats_label = QLabel("📊 الإحصائيات")
        stats_label.setProperty("class", "input-label")
        scroll_layout.addWidget(stats_label)
        
        # Create stats grid
        stats_grid = QVBoxLayout()
        stats_grid.setSpacing(8)
        
        self.total_label = self.create_stat_frame("👥 الإجمالي", "0")
        self.present_label = self.create_stat_frame("✅ الحضور", "0")
        self.attendance_rate_label = self.create_stat_frame("📈 النسبة", "0%")
        self.class1_label = self.create_stat_frame("📚 أولى", "0")
        self.class2_label = self.create_stat_frame("📚 ثانية", "0")
        self.class3_label = self.create_stat_frame("📚 ثالثة", "0")
        
        stats_grid.addWidget(self.total_label)
        stats_grid.addWidget(self.present_label)
        stats_grid.addWidget(self.attendance_rate_label)
        stats_grid.addWidget(self.class1_label)
        stats_grid.addWidget(self.class2_label)
        stats_grid.addWidget(self.class3_label)
        scroll_layout.addLayout(stats_grid)
        
        # Exports
        sep3 = QFrame()
        sep3.setFrameShape(QFrame.HLine)
        sep3.setProperty("class", "card-separator")
        scroll_layout.addWidget(sep3)
        
        export_label = QLabel("📤 التصدير")
        export_label.setProperty("class", "input-label")
        scroll_layout.addWidget(export_label)
        
        export_grid = QVBoxLayout()
        export_grid.setSpacing(8)
        
        # All
        row_all = QHBoxLayout()
        self.export_excel_btn = QPushButton("Excel الكل")
        self.export_excel_btn.setProperty("class", "btn-success")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        
        self.export_pdf_btn = QPushButton("PDF الكل")
        self.export_pdf_btn.setProperty("class", "btn-danger")
        self.export_pdf_btn.clicked.connect(self.export_to_pdf)
        
        row_all.addWidget(self.export_excel_btn)
        row_all.addWidget(self.export_pdf_btn)
        export_grid.addLayout(row_all)
        
        # Classes
        for cls_name, cls_label, btn_excel_name, btn_pdf_name in [
            ('الصف الأول', "أولى", "export_class1_excel_btn", "export_class1_pdf_btn"),
            ('الصف الثاني', "ثانية", "export_class2_excel_btn", "export_class2_pdf_btn"),
            ('الصف الثالث', "ثالثة", "export_class3_excel_btn", "export_class3_pdf_btn")
        ]:
            row = QHBoxLayout()
            btn_excel = QPushButton(f"{cls_label} Excel")
            btn_excel.setProperty("class", "default")
            btn_excel.clicked.connect(lambda c=cls_name: self.export_class_to_excel(c))
            setattr(self, btn_excel_name, btn_excel)
            
            btn_pdf = QPushButton(f"{cls_label} PDF")
            btn_pdf.setProperty("class", "btn-purple")
            btn_pdf.clicked.connect(lambda c=cls_name: self.export_class_to_pdf(c))
            setattr(self, btn_pdf_name, btn_pdf)
            
            row.addWidget(btn_excel)
            row.addWidget(btn_pdf)
            export_grid.addLayout(row)
            
        scroll_layout.addLayout(export_grid)
        scroll_layout.addStretch()
        
        scroll.setWidget(scroll_content)
        left_layout.addWidget(scroll)
        
        main_layout.addWidget(left_panel)
        
        # ═══════════════════════════════════════════════════════════════
        # RIGHT PANEL: Table
        # ═══════════════════════════════════════════════════════════════
        right_panel = QFrame()
        right_panel.setProperty("class", "dashboard-card")
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(12)
        right_layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        table_header = QHBoxLayout()
        table_icon = QLabel("📊")
        table_icon.setProperty("class", "card-icon")
        table_title = QLabel("جدول الحضور")
        table_title.setProperty("class", "card-title")
        table_header.addWidget(table_icon)
        table_header.addWidget(table_title)
        
        table_header.addStretch()
        self.report_info_label = QLabel("")
        self.report_info_label.setProperty("class", "stat-subtitle")
        table_header.addWidget(self.report_info_label)
        
        right_layout.addLayout(table_header)
        
        sep4 = QFrame()
        sep4.setFrameShape(QFrame.HLine)
        sep4.setProperty("class", "card-separator")
        right_layout.addWidget(sep4)
        
        # Table
        self.attendance_table = QTableWidget()
        self.attendance_table.setProperty("class", "dashboard-table")
        self.setup_attendance_table()
        
        right_layout.addWidget(self.attendance_table)
        main_layout.addWidget(right_panel, 1)
        
        self.setLayout(main_layout)
    
    def create_stat_frame(self, label_text, value_text):
        """Create a stat frame widget"""
        frame = QFrame()
        frame.setStyleSheet("background: #0f0f1a; border-radius: 6px;")
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(12, 8, 12, 8)
        
        lbl = QLabel(label_text)
        lbl.setProperty("class", "stat-subtitle")
        
        val = QLabel(value_text)
        val.setStyleSheet("font-weight: 600; color: #e2e8f0;")
        
        layout.addWidget(lbl)
        layout.addStretch()
        layout.addWidget(val)
        
        return frame
    
    def update_stat_frame(self, frame, value):
        """Update stat frame value"""
        labels = frame.findChildren(QLabel)
        if len(labels) >= 2:
            labels[1].setText(str(value))
    
    def setup_attendance_table(self):
        """إعداد جدول الحضور"""
        self.attendance_table.setColumnCount(8)
        self.attendance_table.setHorizontalHeaderLabels([
            "الكود", "الاسم", "الصف", "المنطقة", "العمارة", "الشارع", "وقت الحضور", "التاريخ"
        ])
        
        self.attendance_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.attendance_table.setAlternatingRowColors(True)
        self.attendance_table.setSortingEnabled(True)
        
        # ضبط أبعاد الأعمدة
        self.attendance_table.setColumnWidth(0, 80)   # الكود
        self.attendance_table.setColumnWidth(1, 150)  # الاسم
        self.attendance_table.setColumnWidth(2, 100)  # الصف
        self.attendance_table.setColumnWidth(3, 100)  # المنطقة
        self.attendance_table.setColumnWidth(4, 80)   # العمارة
        self.attendance_table.setColumnWidth(5, 120)  # الشارع
        self.attendance_table.setColumnWidth(6, 100)  # وقت الحضور
        self.attendance_table.setColumnWidth(7, 100)  # التاريخ
    
    def load_dates(self):
        """تحميل تواريخ الحضور المسجلة"""
        try:
            dates_info = self.db.get_attendance_dates_with_info()
            
            # إذا كان مختار "أيام الخدمة فقط"
            if self.service_days_only.isChecked():
                dates_info = [date for date in dates_info if date['is_service_day']]
            
            self.date_selector.clear()
            for date_info in dates_info:
                self.date_selector.addItem(date_info['display'], date_info['date'])
            
            if dates_info:
                self.date_selector.setCurrentIndex(0)
                self.update_date_info(dates_info[0])
            else:
                self.date_info_label.setText("لا توجد تواريخ مسجلة")
                
        except Exception as e:
            QMessageBox.warning(self, "تحذير", f"خطأ في تحميل التواريخ: {str(e)}")
    
    def update_date_info(self, date_info):
        """تحديث معلومات التاريخ المحدد"""
        info_text = f"التاريخ: {date_info['date']} | اليوم: {date_info['day_name']}"
        if date_info['is_service_day']:
            info_text += " | ✅ يوم الخدمة"
        else:
            info_text += " | ⚠️ ليس يوم خدمة"
        
        self.date_info_label.setText(info_text)
    
    def load_attendance_data(self):
        """تحميل بيانات الحضور للتاريخ المحدد"""
        try:
            if self.date_selector.currentIndex() < 0:
                return
                
            selected_date = self.date_selector.currentData()
            if not selected_date:
                display_text = self.date_selector.currentText()
                selected_date = display_text.split(' ')[0]
            
            self.current_date = selected_date
            
            # تحديث معلومات التاريخ
            date_obj = datetime.strptime(selected_date, "%Y-%m-%d")
            day_name = date_obj.strftime("%A")
            arabic_day = self.db.arabic_days.get(day_name, day_name)
            settings = self.db.get_settings()
            is_service_day = (day_name == settings.get('service_day', 'Thursday'))
            
            self.update_date_info({
                'date': selected_date,
                'day_name': arabic_day,
                'is_service_day': is_service_day
            })
            
            # الحصول على بيانات الحضور
            data = self.db.load_data()
            attendance_data = data.get('attendance', {}).get(selected_date, {})
            
            # جمع بيانات الأطفال الحاضرين
            self.current_attendance_data = []
            for code, arrival_time in attendance_data.items():
                child = self.db.get_child_by_code(code)
                if child:
                    attendance_record = child.copy()
                    attendance_record['arrival_time'] = arrival_time
                    attendance_record['attendance_date'] = selected_date
                    self.current_attendance_data.append(attendance_record)
            
            # تحديث الجدول
            self.update_attendance_table()
            
            # تحديث الإحصائيات
            self.update_stats(selected_date)
            
            # تحديث معلومات التقرير
            self.report_info_label.setText(f"تقرير الحضور الشامل - {selected_date} | عدد الحاضرين: {len(self.current_attendance_data)}")
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"خطأ في تحميل بيانات الحضور: {str(e)}")
    
    def update_attendance_table(self):
        """تحديث جدول الحضور"""
        try:
            self.attendance_table.setSortingEnabled(False)
            self.attendance_table.setRowCount(0)
            
            for record in self.current_attendance_data:
                row = self.attendance_table.rowCount()
                self.attendance_table.insertRow(row)
                
                self.attendance_table.setItem(row, 0, QTableWidgetItem(str(record.get('code', 'غير محدد'))))
                self.attendance_table.setItem(row, 1, QTableWidgetItem(record.get('name', 'غير محدد')))
                self.attendance_table.setItem(row, 2, QTableWidgetItem(record.get('class', 'غير محدد')))
                self.attendance_table.setItem(row, 3, QTableWidgetItem(record.get('region', 'غير محدد')))
                self.attendance_table.setItem(row, 4, QTableWidgetItem(str(record.get('عماره', 'غير محدد'))))
                self.attendance_table.setItem(row, 5, QTableWidgetItem(record.get('شارع', 'غير محدد')))
                self.attendance_table.setItem(row, 6, QTableWidgetItem(record.get('arrival_time', 'غير محدد')))
                self.attendance_table.setItem(row, 7, QTableWidgetItem(record.get('attendance_date', 'غير محدد')))
            
            self.attendance_table.setSortingEnabled(True)
            
            if len(self.current_attendance_data) == 0:
                self.attendance_table.setRowCount(1)
                item = QTableWidgetItem("⚠️ لا يوجد حضور مسجل لهذا اليوم")
                item.setTextAlignment(Qt.AlignCenter)
                self.attendance_table.setItem(0, 0, item)
                self.attendance_table.setSpan(0, 0, 1, 8)
                
        except Exception as e:
            print(f"Error updating attendance table: {str(e)}")
    
    def update_stats(self, date):
        """تحديث الإحصائيات"""
        try:
            stats = self.db.get_attendance_stats_by_date(date)
            if not stats:
                self.update_stat_frame(self.total_label, "0")
                self.update_stat_frame(self.present_label, "0")
                self.update_stat_frame(self.attendance_rate_label, "0%")
                self.update_stat_frame(self.class1_label, "0")
                self.update_stat_frame(self.class2_label, "0")
                self.update_stat_frame(self.class3_label, "0")
                return
            
            self.update_stat_frame(self.total_label, str(stats['total']))
            self.update_stat_frame(self.present_label, str(stats['present']))
            self.update_stat_frame(self.attendance_rate_label, f"{stats['attendance_rate']:.1f}%")
            
            class_stats = stats.get('classes', {})
            self.update_stat_frame(self.class1_label, str(class_stats.get('الصف الأول', {}).get('present', 0)))
            self.update_stat_frame(self.class2_label, str(class_stats.get('الصف الثاني', {}).get('present', 0)))
            self.update_stat_frame(self.class3_label, str(class_stats.get('الصف الثالث', {}).get('present', 0)))
            
        except Exception as e:
            print(f"Error updating stats: {str(e)}")
    
    def refresh_data(self):
        """تحديث البيانات"""
        self.load_dates()
        if self.date_selector.currentIndex() >= 0:
            self.load_attendance_data()
        QMessageBox.information(self, "تم التحديث", "تم تحديث بيانات الحضور بنجاح")
    
    def get_class_data(self, class_name):
        """الحصول على بيانات الحضور لصف معين"""
        if not self.current_attendance_data:
            return []
        
        return [record for record in self.current_attendance_data if record.get('class') == class_name]
    
    def export_class_to_excel(self, class_name):
        """تصدير تقرير الحضور لصف معين إلى Excel"""
        class_data = self.get_class_data(class_name)
        
        if not class_data:
            QMessageBox.information(self, "معلومة", f"لا توجد بيانات حضور للصف {class_name}")
            return
        
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, f"حفظ تقرير الحضور - {class_name}", 
                f"تقرير_الحضور_{class_name}_{self.current_date}.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                # إنشاء DataFrame للبيانات
                data = []
                for record in class_data:
                    data.append({
                        'الكود': record.get('code', 'غير محدد'),
                        'الاسم': record.get('name', 'غير محدد'),
                        'الصف': record.get('class', 'غير محدد'),
                        'المنطقة': record.get('region', 'غير محدد'),
                        'العمارة': record.get('عماره', 'غير محدد'),
                        'الشارع': record.get('شارع', 'غير محدد'),
                        'الدور': record.get('دور', 'غير محدد'),
                        'الشقة': record.get('شقه', 'غير محدد'),
                        'موبيل الولد': record.get('موبيل الولد', 'غير محدد'),
                        'موبايل الأب': record.get('موبايل الاب', 'غير محدد'),
                        'موبايل الأم': record.get('موبايل الام', 'غير محدد'),
                        'تليفون': record.get('تليفون', 'غير محدد'),
                        'وقت الحضور': record.get('arrival_time', 'غير محدد'),
                        'تاريخ الحضور': record.get('attendance_date', 'غير محدد')
                    })
                
                # إنشاء ملف Excel مع تنسيق محسن
                wb = Workbook()
                ws = wb.active
                ws.title = f"تقرير الحضور - {class_name}"
                
                # إضافة العناوين
                titles = [
                    f"تقرير الحضور - {class_name}",
                    f"تاريخ التقرير: {self.current_date}",
                    f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    f"عدد الحاضرين: {len(class_data)}"
                ]
                
                for i, title in enumerate(titles, 1):
                    ws[f'A{i}'] = title
                    ws[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                    ws[f'A{i}'].alignment = Alignment(horizontal='center')
                
                # دمج الخلايا للعناوين
                ws.merge_cells('A1:N1')
                ws.merge_cells('A2:N2')
                ws.merge_cells('A3:N3')
                ws.merge_cells('A4:N4')
                
                # إضافة رؤوس الأعمدة
                headers = ['الكود', 'الاسم', 'الصف', 'المنطقة', 'العمارة', 'الشارع', 'الدور', 'الشقة',
                          'موبيل الولد', 'موبايل الأب', 'موبايل الأم', 'تليفون', 'وقت الحضور', 'تاريخ الحضور']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=6, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                
                # إضافة البيانات
                for row, record in enumerate(data, 7):
                    for col, key in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=record.get(key, ''))
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='thin'))
                
                # ضبط أبعاد الأعمدة
                column_widths = {
                    'A': 12, 'B': 25, 'C': 15, 'D': 15, 'E': 12, 'F': 20, 'G': 10, 'H': 10,
                    'I': 15, 'J': 15, 'K': 15, 'L': 15, 'M': 12, 'N': 12
                }
                
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # حفظ الملف
                wb.save(file_path)
                
                QMessageBox.information(self, "تم التصدير", 
                                      f"تم تصدير تقرير الحضور للصف {class_name} إلى Excel بنجاح!\n\n"
                                      f"الملف: {file_path}\n"
                                      f"عدد السجلات: {len(class_data)}")
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"خطأ في التصدير إلى Excel: {str(e)}")
    
    def export_class_to_pdf(self, class_name):
        """تصدير تقرير الحضور لصف معين إلى PDF"""
        class_data = self.get_class_data(class_name)
        
        if not class_data:
            QMessageBox.information(self, "معلومة", f"لا توجد بيانات حضور للصف {class_name}")
            return
        
        try:
            from PyQt5.QtWidgets import QFileDialog
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, f"حفظ تقرير PDF - {class_name}", 
                f"تقرير_الحضور_{class_name}_{self.current_date}.pdf", 
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
            
            # البحث عن خطوط عربية
            arabic_font_paths = [
                "fonts/arial.ttf",
                "fonts/tahoma.ttf",
                "fonts/arabic_font.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/tahoma.ttf"
            ]
            
            arabic_font_name = None
            for font_path in arabic_font_paths:
                if os.path.exists(font_path):
                    try:
                        font_name = os.path.splitext(os.path.basename(font_path))[0]
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                        arabic_font_name = font_name
                        break
                    except:
                        continue
            
            if not arabic_font_name:
                arabic_font_name = "Helvetica"
            
            # إنشاء مستند PDF
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4),
                                  rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()
            
            # إضافة عنوان
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=arabic_font_name,
                fontSize=16,
                spaceAfter=20,
                alignment=1,
                textColor=colors.HexColor('#2C3E50')
            )
            
            title_text = self.reshape_arabic(f"تقرير الحضور - {class_name}")
            elements.append(Paragraph(title_text, title_style))
            
            # معلومات التقرير
            info_style = ParagraphStyle(
                'CustomInfo',
                parent=styles['Normal'],
                fontName=arabic_font_name,
                fontSize=10,
                spaceAfter=8,
                alignment=1
            )
            
            info_texts = [
                f"تاريخ التقرير: {self.current_date}",
                f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                f"عدد الحاضرين: {len(class_data)}"
            ]
            
            for text in info_texts:
                elements.append(Paragraph(self.reshape_arabic(text), info_style))
            
            elements.append(Spacer(1, 20))
            
            # رؤوس الأعمدة (من اليمين لليسار)
            headers = [
                'تاريخ الحضور', 'وقت الحضور', 'تليفون', 'موبايل الأم', 'موبايل الأب', 
                'موبيل الولد', 'الشقة', 'الدور', 'الشارع', 'العمارة', 
                'المنطقة', 'الصف', 'الاسم', 'الكود'
            ]
            
            # إعداد بيانات الجدول
            table_data = [headers]
            
            for record in class_data:
                table_data.append([
                    self.reshape_arabic(record.get('attendance_date', 'غير محدد')),
                    self.reshape_arabic(record.get('arrival_time', 'غير محدد')),
                    self.reshape_arabic(record.get('تليفون', 'غير محدد')),
                    self.reshape_arabic(record.get('موبايل الام', 'غير محدد')),
                    self.reshape_arabic(record.get('موبايل الاب', 'غير محدد')),
                    self.reshape_arabic(record.get('موبيل الولد', 'غير محدد')),
                    self.reshape_arabic(record.get('شقه', 'غير محدد')),
                    self.reshape_arabic(record.get('دور', 'غير محدد')),
                    self.reshape_arabic(record.get('شارع', 'غير محدد')),
                    self.reshape_arabic(record.get('عماره', 'غير محدد')),
                    self.reshape_arabic(record.get('region', 'غير محدد')),
                    self.reshape_arabic(record.get('class', 'غير محدد')),
                    self.reshape_arabic(record.get('name', 'غير محدد')),
                    self.reshape_arabic(record.get('code', 'غير محدد'))
                ])
            
            # إنشاء الجدول
            table = Table(table_data, repeatRows=1)
            
            # تنسيق الجدول
            table.setStyle(TableStyle([
                # تنسيق رأس الجدول
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), arabic_font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # تنسيق بيانات الجدول
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), arabic_font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # الحدود
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                
                # التفاف النص
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ]))
            
            # ضبط أبعاد الأعمدة
            col_widths = [50, 50, 50, 50, 50, 50, 40, 40, 50, 40, 40, 50, 50, 40]
            table._argW = col_widths
            
            elements.append(table)
            
            # بناء PDF
            doc.build(elements)
            QMessageBox.information(self, "تم التصدير", f"تم تصدير تقرير PDF للصف {class_name} بنجاح!")
            
        except ImportError as e:
            QMessageBox.critical(self, "خطأ", 
                               f"المكتبات المطلوبة غير مثبتة!\n"
                               f"يرجى تثبيت المكتبات التالية:\n"
                               f"pip install reportlab arabic-reshaper python-bidi")
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء التصدير إلى PDF: {str(e)}")
    
    # الدوال الأصلية تبقى كما هي مع تعديلات بسيطة
    def export_to_excel(self):
        """تصدير تقرير الحضور إلى Excel"""
        if not self.current_attendance_data:
            QMessageBox.information(self, "معلومة", "لا توجد بيانات حضور للتصدير")
            return
        
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ تقرير الحضور", 
                f"تقرير_الحضور_{self.current_date}.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                # إنشاء DataFrame للبيانات
                data = []
                for record in self.current_attendance_data:
                    data.append({
                        'الكود': record.get('code', 'غير محدد'),
                        'الاسم': record.get('name', 'غير محدد'),
                        'الصف': record.get('class', 'غير محدد'),
                        'المنطقة': record.get('region', 'غير محدد'),
                        'العمارة': record.get('عماره', 'غير محدد'),
                        'الشارع': record.get('شارع', 'غير محدد'),
                        'الدور': record.get('دور', 'غير محدد'),
                        'الشقة': record.get('شقه', 'غير محدد'),
                        'موبيل الولد': record.get('موبيل الولد', 'غير محدد'),
                        'موبايل الأب': record.get('موبايل الاب', 'غير محدد'),
                        'موبايل الأم': record.get('موبايل الام', 'غير محدد'),
                        'تليفون': record.get('تليفون', 'غير محدد'),
                        'وقت الحضور': record.get('arrival_time', 'غير محدد'),
                        'تاريخ الحضور': record.get('attendance_date', 'غير محدد')
                    })
                
                # إنشاء ملف Excel مع تنسيق محسن
                wb = Workbook()
                ws = wb.active
                ws.title = "تقرير الحضور"
                
                # إضافة العناوين
                titles = [
                    f"تقرير الحضور الشامل",
                    f"تاريخ التقرير: {self.current_date}",
                    f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    f"عدد الحاضرين: {len(self.current_attendance_data)}"
                ]
                
                for i, title in enumerate(titles, 1):
                    ws[f'A{i}'] = title
                    ws[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                    ws[f'A{i}'].alignment = Alignment(horizontal='center')
                
                # دمج الخلايا للعناوين
                ws.merge_cells('A1:N1')
                ws.merge_cells('A2:N2')
                ws.merge_cells('A3:N3')
                ws.merge_cells('A4:N4')
                
                # إضافة رؤوس الأعمدة
                headers = ['الكود', 'الاسم', 'الصف', 'المنطقة', 'العمارة', 'الشارع', 'الدور', 'الشقة',
                          'موبيل الولد', 'موبايل الأب', 'موبايل الأم', 'تليفون', 'وقت الحضور', 'تاريخ الحضور']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=6, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                
                # إضافة البيانات
                for row, record in enumerate(data, 7):
                    for col, key in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=record.get(key, ''))
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='thin'))
                
                # ضبط أبعاد الأعمدة
                column_widths = {
                    'A': 12, 'B': 25, 'C': 15, 'D': 15, 'E': 12, 'F': 20, 'G': 10, 'H': 10,
                    'I': 15, 'J': 15, 'K': 15, 'L': 15, 'M': 12, 'N': 12
                }
                
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # حفظ الملف
                wb.save(file_path)
                
                QMessageBox.information(self, "تم التصدير", 
                                      f"تم تصدير تقرير الحضور إلى Excel بنجاح!\n\n"
                                      f"الملف: {file_path}\n"
                                      f"عدد السجلات: {len(self.current_attendance_data)}")
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"خطأ في التصدير إلى Excel: {str(e)}")
    
    def export_to_pdf(self):
        """تصدير تقرير الحضور إلى PDF"""
        if not self.current_attendance_data:
            QMessageBox.information(self, "معلومة", "لا توجد بيانات حضور للتصدير")
            return
        
        try:
            from PyQt5.QtWidgets import QFileDialog
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ تقرير PDF", 
                f"تقرير_الحضور_{self.current_date}.pdf", 
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
            
            # البحث عن خطوط عربية
            arabic_font_paths = [
                "fonts/arial.ttf",
                "fonts/tahoma.ttf",
                "fonts/arabic_font.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/tahoma.ttf"
            ]
            
            arabic_font_name = None
            for font_path in arabic_font_paths:
                if os.path.exists(font_path):
                    try:
                        font_name = os.path.splitext(os.path.basename(font_path))[0]
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                        arabic_font_name = font_name
                        break
                    except:
                        continue
            
            if not arabic_font_name:
                arabic_font_name = "Helvetica"
            
            # إنشاء مستند PDF
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4),
                                  rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()
            
            # إضافة عنوان رئيسي
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=arabic_font_name,
                fontSize=16,
                spaceAfter=20,
                alignment=1,
                textColor=colors.HexColor('#2C3E50')
            )
            
            title_text = self.reshape_arabic("تقرير الحضور الشامل")
            elements.append(Paragraph(title_text, title_style))
            
            # معلومات التقرير
            info_style = ParagraphStyle(
                'CustomInfo',
                parent=styles['Normal'],
                fontName=arabic_font_name,
                fontSize=10,
                spaceAfter=8,
                alignment=1
            )
            
            info_texts = [
                f"تاريخ التقرير: {self.current_date}",
                f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                f"عدد الحاضرين: {len(self.current_attendance_data)}"
            ]
            
            for text in info_texts:
                elements.append(Paragraph(self.reshape_arabic(text), info_style))
            
            elements.append(Spacer(1, 20))
            
            # رؤوس الأعمدة (من اليمين لليسار)
            headers = [
                'تاريخ الحضور', 'وقت الحضور', 'تليفون', 'موبايل الأم', 'موبايل الأب', 
                'موبيل الولد', 'الشقة', 'الدور', 'الشارع', 'العمارة', 
                'المنطقة', 'الصف', 'الاسم', 'الكود'
            ]
            
            # إعداد بيانات الجدول
            table_data = [headers]
            
            for record in self.current_attendance_data:
                table_data.append([
                    self.reshape_arabic(record.get('attendance_date', 'غير محدد')),
                    self.reshape_arabic(record.get('arrival_time', 'غير محدد')),
                    self.reshape_arabic(record.get('تليفون', 'غير محدد')),
                    self.reshape_arabic(record.get('موبايل الام', 'غير محدد')),
                    self.reshape_arabic(record.get('موبايل الاب', 'غير محدد')),
                    self.reshape_arabic(record.get('موبيل الولد', 'غير محدد')),
                    self.reshape_arabic(record.get('شقه', 'غير محدد')),
                    self.reshape_arabic(record.get('دور', 'غير محدد')),
                    self.reshape_arabic(record.get('شارع', 'غير محدد')),
                    self.reshape_arabic(record.get('عماره', 'غير محدد')),
                    self.reshape_arabic(record.get('region', 'غير محدد')),
                    self.reshape_arabic(record.get('class', 'غير محدد')),
                    self.reshape_arabic(record.get('name', 'غير محدد')),
                    self.reshape_arabic(record.get('code', 'غير محدد'))
                ])
            
            # إنشاء الجدول
            table = Table(table_data, repeatRows=1)
            
            # تنسيق الجدول
            table.setStyle(TableStyle([
                # تنسيق رأس الجدول
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), arabic_font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # تنسيق بيانات الجدول
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), arabic_font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # الحدود
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                
                # التفاف النص
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ]))
            
            # ضبط أبعاد الأعمدة
            col_widths = [50, 50, 50, 50, 50, 50, 40, 40, 50, 40, 40, 50, 50, 40]
            table._argW = col_widths
            
            elements.append(table)
            
            # بناء PDF
            doc.build(elements)
            QMessageBox.information(self, "تم التصدير", "تم تصدير تقرير PDF بنجاح!")
            
        except ImportError as e:
            QMessageBox.critical(self, "خطأ", 
                               f"المكتبات المطلوبة غير مثبتة!\n"
                               f"يرجى تثبيت المكتبات التالية:\n"
                               f"pip install reportlab arabic-reshaper python-bidi")
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء التصدير إلى PDF: {str(e)}")
    
    def reshape_arabic(self, text):
        """إعادة تشكيل النص العربي للعرض الصحيح في PDF"""
        try:
            if not text or text == 'غير محدد':
                return 'غير محدد'
            
            # إعادة تشكيل النص العربي
            reshaped_text = arabic_reshaper.reshape(str(text))
            # تحويل النص إلى تنسيق من اليمين لليسار
            bidi_text = get_display(reshaped_text)
            return bidi_text
        except:
            # في حالة حدوث خطأ، إرجاع النص الأصلي
            return str(text)
