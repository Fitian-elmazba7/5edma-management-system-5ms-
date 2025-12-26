# excel_handler.py
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from .coptic_excel_parser import CopticExcelParser

class ExcelHandler:
    def __init__(self):
        self.parser = CopticExcelParser()
    
    def import_from_excel(self, file_path, db, incremental=False):
        """استيراد البيانات من ملف Excel إلى قاعدة البيانات مع دعم الاستيراد التزايدي"""
        try:
            # الحصول على الأكواد الموجودة مسبقاً إذا كان الاستيراد تزايدياً
            existing_codes = None
            if incremental:
                existing_children = db.get_all_children()
                existing_codes = {str(child.get('code', '')).strip() for child in existing_children if child.get('code')}
                print(f"تم العثور على {len(existing_codes)} طفل موجود مسبقاً في قاعدة البيانات")
            
            # تحليل ملف Excel مع التحقق من الأكواد المكررة
            children_data = self.parser.parse_excel_file(file_path, existing_codes)
            
            # التحقق من صحة البيانات
            valid_children, issues = self.parser.validate_data(children_data)
            
            # استيراد البيانات إلى قاعدة البيانات
            imported_count = db.import_children(valid_children)
            
            result_info = {
                'issues': issues, 
                'valid_count': len(valid_children),
                'imported_count': imported_count,
                'skipped_count': len(existing_codes) if existing_codes else 0
            }
            
            return imported_count, result_info
            
        except Exception as e:
            raise Exception(f"خطأ في استيراد ملف Excel: {str(e)}")
    
    def analyze_excel_file(self, file_path):
        """تحليل ملف Excel وعرض هيكله"""
        try:
            # قراءة ملف Excel
            df = pd.read_excel(file_path)
            
            # تحليل الهيكل
            analysis = {
                'file_info': {
                    'file_name': os.path.basename(file_path),
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'columns': list(df.columns),
                    'required_columns_found': len([col for col in self.parser.primary_columns if col in df.columns])
                },
                'sample_data': {}
            }
            
            # عينة من البيانات (الصف الأول)
            if len(df) > 0:
                first_row = df.iloc[0]
                for col in df.columns:
                    analysis['sample_data'][col] = str(first_row[col]) if pd.notna(first_row[col]) else 'فارغ'
            
            return analysis
            
        except Exception as e:
            raise Exception(f"خطأ في تحليل ملف Excel: {str(e)}")
    
    def export_absence_report(self, absent_children, file_path, date, report_title):
        """تصدير تقرير غياب مفصل مع تنسيق محسن مشابه لتقرير الحضور"""
        try:
            if not absent_children:
                raise Exception("لا توجد بيانات غياب للتصدير")
            
            # إنشاء DataFrame للبيانات
            data = []
            for child in absent_children:
                data.append({
                    'الكود': child.get('code', 'غير محدد'),
                    'الاسم': child.get('name', 'غير محدد'),
                    'الصف': child.get('class', 'غير محدد'),
                    'المنطقة': child.get('region', 'غير محدد'),
                    'العمارة': child.get('عماره', 'غير محدد'),
                    'الشارع': child.get('شارع', 'غير محدد'),
                    'الدور': child.get('دور', 'غير محدد'),
                    'الشقة': child.get('شقه', 'غير محدد'),
                    'موبيل الولد': child.get('موبيل الولد', 'غير محدد'),
                    'موبايل الأب': child.get('موبايل الاب', 'غير محدد'),
                    'موبايل الأم': child.get('موبايل الام', 'غير محدد'),
                    'تليفون': child.get('تليفون', 'غير محدد'),
                    'المدرسة': child.get('المدرسه', 'غير محدد'),
                    'ملاحظات': child.get('ملاحظات', 'غير محدد'),
                    'تاريخ الغياب': date
                })
            
            # إنشاء ملف Excel مع تنسيق محسن
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "تقرير الغياب"
            
            # إضافة العناوين
            titles = [
                f"تقرير {report_title}",
                f"تاريخ الغياب: {date}",
                f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                f"عدد الأطفال الغائبين: {len(absent_children)}"
            ]
            
            for i, title in enumerate(titles, 1):
                ws[f'A{i}'] = title
                ws[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                ws[f'A{i}'].alignment = Alignment(horizontal='center')
            
            # دمج الخلايا للعناوين
            ws.merge_cells('A1:O1')
            ws.merge_cells('A2:O2')
            ws.merge_cells('A3:O3')
            ws.merge_cells('A4:O4')
            
            # إضافة رؤوس الأعمدة
            headers = ['الكود', 'الاسم', 'الصف', 'المنطقة', 'العمارة', 'الشارع', 'الدور', 'الشقة',
                      'موبيل الولد', 'موبايل الأب', 'موبايل الأم', 'تليفون', 'المدرسة', 'ملاحظات', 'تاريخ الغياب']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # إضافة البيانات
            for row, record in enumerate(data, 7):
                for col, key in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=record.get(key, ''))
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
            
            # ضبط أبعاد الأعمدة
            column_widths = {
                'A': 12, 'B': 25, 'C': 15, 'D': 15, 'E': 12, 'F': 20, 'G': 10, 'H': 10,
                'I': 15, 'J': 15, 'K': 15, 'L': 15, 'M': 20, 'N': 30, 'O': 15
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # حفظ الملف
            wb.save(file_path)
            
            print(f"تم تصدير تقرير الغياب المحسن إلى: {file_path}")
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير تقرير الغياب: {str(e)}")
    
    def export_to_excel(self, children, file_path):
        """تصدير جميع بيانات الأطفال إلى ملف Excel"""
        try:
            data = []
            for child in children:
                data.append({
                    'الكود': child.get('code', 'غير محدد'),
                    'الاسم': child.get('name', 'غير محدد'),
                    'الصف': child.get('class', 'غير محدد'),
                    'المنطقة': child.get('region', 'غير محدد'),
                    'العمارة': child.get('عماره', 'غير محدد'),
                    'الشارع': child.get('شارع', 'غير محدد'),
                    'الدور': child.get('دور', 'غير محدد'),
                    'الشقة': child.get('شقه', 'غير محدد'),
                    'موبيل الولد': child.get('موبيل الولد', 'غير محدد'),
                    'موبايل الأب': child.get('موبايل الاب', 'غير محدد'),
                    'موبايل الأم': child.get('موبايل الام', 'غير محدد'),
                    'تليفون': child.get('تليفون', 'غير محدد'),
                    'المدرسة': child.get('المدرسه', 'غير محدد'),
                    'ملاحظات': child.get('ملاحظات', 'غير محدد')
                })
            
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False)
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير البيانات: {str(e)}")

    def export_early_arrival_report(self, early_arrivals, file_path, date, threshold_time):
        """تصدير تقرير الحضور المبكر المحسن"""
        try:
            # إنشاء ملف Excel مع تنسيق محسن
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "الحضور المبكر"
            
            # إضافة العناوين
            titles = [
                f"تقرير الحضور المبكر",
                f"تاريخ التقرير: {date}",
                f"الوقت المحدد للحضور المبكر: {threshold_time}",
                f"عدد الأطفال في التقرير: {len(early_arrivals)}"
            ]
            
            for i, title in enumerate(titles, 1):
                ws[f'A{i}'] = title
                ws[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                ws[f'A{i}'].alignment = Alignment(horizontal='center')
            
            # دمج الخلايا للعناوين
            ws.merge_cells('A1:G1')
            ws.merge_cells('A2:G2')
            ws.merge_cells('A3:G3')
            ws.merge_cells('A4:G4')
            
            # إضافة رؤوس الأعمدة
            headers = ['الكود', 'الاسم', 'الصف', 'المنطقة', 'وقت الحضور', 'التاريخ', 'ملاحظات']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # إضافة البيانات
            for row, record in enumerate(early_arrivals, 7):
                ws.cell(row=row, column=1, value=record.get('code', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=2, value=record.get('name', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=3, value=record.get('class', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=4, value=record.get('region', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=record.get('arrival_time', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=6, value=record.get('attendance_date', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=7, value=record.get('notes', '')).alignment = Alignment(horizontal='center')
            
            # ضبط أبعاد الأعمدة
            column_widths = {'A': 12, 'B': 25, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 30}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # حفظ الملف
            wb.save(file_path)
            
            print(f"تم تصدير تقرير الحضور المبكر المحسن إلى: {file_path}")
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير تقرير الحضور المبكر: {str(e)}")

    def export_server_assignment_report(self, assignments_data, file_path, date):
        """تصدير تقرير توزيع المتابعة على الخدام"""
        try:
            # إنشاء ملف Excel مع تنسيق محسن
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "توزيع المتابعة"
            
            # إضافة العناوين
            titles = [
                f"تقرير توزيع متابعة الغياب على الخدام",
                f"تاريخ الغياب: {date}",
                f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                f"عدد الأطفال: {len(assignments_data)}"
            ]
            
            for i, title in enumerate(titles, 1):
                ws[f'A{i}'] = title
                ws[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                ws[f'A{i}'].alignment = Alignment(horizontal='center')
            
            # دمج الخلايا للعناوين
            ws.merge_cells('A1:J1')
            ws.merge_cells('A2:J2')
            ws.merge_cells('A3:J3')
            ws.merge_cells('A4:J4')
            
            # إضافة رؤوس الأعمدة
            headers = ['الصف', 'الخادم', 'الكود', 'الاسم', 'العنوان', 'هاتف الطفل', 'هاتف الأب', 'هاتف الأم', 'تليفون المنزل', 'تاريخ الغياب']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # إضافة البيانات
            for row, assignment in enumerate(assignments_data, 7):
                ws.cell(row=row, column=1, value=assignment.get('class', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=2, value=assignment.get('server', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=3, value=assignment.get('code', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=4, value=assignment.get('name', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=assignment.get('address', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=6, value=assignment.get('child_phone', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=7, value=assignment.get('father_phone', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=8, value=assignment.get('mother_phone', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=9, value=assignment.get('home_phone', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=10, value=assignment.get('absence_date', '')).alignment = Alignment(horizontal='center')
            
            # ضبط أبعاد الأعمدة
            column_widths = {
                'A': 15, 'B': 20, 'C': 12, 'D': 25, 'E': 30, 
                'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 15
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # حفظ الملف
            wb.save(file_path)
            
            print(f"تم تصدير تقرير توزيع المتابعة المحسن إلى: {file_path}")
            
        except Exception as e:
            raise Exception(f"خطأ في تصدير تقرير توزيع المتابعة: {str(e)}")