from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                             QTableWidget, QTableWidgetItem, QHeaderView,
                             QPushButton, QSpinBox, QGroupBox, QMessageBox,
                             QTabWidget, QProgressBar, QSplitter, QComboBox,
                             QFormLayout, QFrame, QWidget)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon, QFont
import random
from datetime import datetime
import os

class ServerAssignmentDialog(QDialog):
    def __init__(self, absent_children, date, parent=None):
        super().__init__(parent)
        self.absent_children = absent_children
        self.date = date
        self.setWindowTitle(f"توزيع المتابعة على الخدام - {date}")
        self.setModal(True)
        self.setMinimumSize(1200, 800)
        self.server_assignments = {}
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # عنوان النافذة
        title_label = QLabel(f"🎯 توزيع متابعة الغياب على الخدام - {self.date}")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setProperty("class", "page-title")

        
        # معلومات الإحصائيات
        stats_frame = QFrame()
        stats_frame.setProperty("class", "stats-container")

        stats_layout = QHBoxLayout()
        
        total_absent = len(self.absent_children)
        class1_count = len([c for c in self.absent_children if c.get('class') == 'الصف الأول'])
        class2_count = len([c for c in self.absent_children if c.get('class') == 'الصف الثاني'])
        class3_count = len([c for c in self.absent_children if c.get('class') == 'الصف الثالث'])
        
        stats_labels = [
            f"👥 إجمالي الغياب: {total_absent}",
            f"📚 الصف الأول: {class1_count}",
            f"📚 الصف الثاني: {class2_count}",
            f"📚 الصف الثالث: {class3_count}"
        ]
        
        for text in stats_labels:
            label = QLabel(text)
            label.setProperty("class", "status-good")

            stats_layout.addWidget(label)
        
        stats_frame.setLayout(stats_layout)
        
        # إعدادات التوزيع
        settings_group = QGroupBox("⚙️ إعدادات التوزيع")
        settings_layout = QFormLayout()
        
        # عدد الخدام لكل صف
        self.servers_class1 = QSpinBox()
        self.servers_class1.setRange(1, 10)
        self.servers_class1.setValue(10)
        self.servers_class1.valueChanged.connect(self.update_distribution)
        
        self.servers_class2 = QSpinBox()
        self.servers_class2.setRange(1, 10)
        self.servers_class2.setValue(9)
        self.servers_class2.valueChanged.connect(self.update_distribution)
        
        self.servers_class3 = QSpinBox()
        self.servers_class3.setRange(1, 10)
        self.servers_class3.setValue(10)
        self.servers_class3.valueChanged.connect(self.update_distribution)
        
        # أسماء الخدام الافتراضية
        self.default_servers = [
            "الخادم الأول", "الخادم الثاني", "الخادم الثالث", "الخادم الرابع",
            "الخادم الخامس", "الخادم السادس", "الخادم السابع", "الخادم الثامن",
            "الخادم التاسع", "الخادم العاشر"
        ]
        
        settings_layout.addRow("عدد خدام الصف الأول:", self.servers_class1)
        settings_layout.addRow("عدد خدام الصف الثاني:", self.servers_class2)
        settings_layout.addRow("عدد خدام الصف الثالث:", self.servers_class3)
        
        settings_group.setLayout(settings_layout)
        
        # أزرار التحكم
        control_layout = QHBoxLayout()
        
        self.distribute_btn = QPushButton("🔄 توزيع عشوائي")
        self.distribute_btn.clicked.connect(self.distribute_children)
        
        # مجموعة تصدير Excel
        excel_export_layout = QVBoxLayout()
        excel_export_layout.addWidget(QLabel("📊 تصدير Excel:"))
        
        excel_buttons_layout = QHBoxLayout()
        self.export_all_excel_btn = QPushButton("📊 تصدير الكل Excel")
        self.export_all_excel_btn.clicked.connect(self.export_all_assignments_excel)
        
        self.export_class1_excel_btn = QPushButton("📊 الصف الأول Excel")
        self.export_class1_excel_btn.clicked.connect(lambda: self.export_class_assignment_excel('الصف الأول'))
        
        self.export_class2_excel_btn = QPushButton("📊 الصف الثاني Excel")
        self.export_class2_excel_btn.clicked.connect(lambda: self.export_class_assignment_excel('الصف الثاني'))
        
        self.export_class3_excel_btn = QPushButton("📊 الصف الثالث Excel")
        self.export_class3_excel_btn.clicked.connect(lambda: self.export_class_assignment_excel('الصف الثالث'))
        
        excel_buttons_layout.addWidget(self.export_all_excel_btn)
        excel_buttons_layout.addWidget(self.export_class1_excel_btn)
        excel_buttons_layout.addWidget(self.export_class2_excel_btn)
        excel_buttons_layout.addWidget(self.export_class3_excel_btn)
        excel_export_layout.addLayout(excel_buttons_layout)
        
        # مجموعة تصدير PDF
        pdf_export_layout = QVBoxLayout()
        pdf_export_layout.addWidget(QLabel("📄 تصدير PDF:"))
        
        pdf_buttons_layout = QHBoxLayout()
        self.export_all_pdf_btn = QPushButton("📄 تصدير الكل PDF")
        self.export_all_pdf_btn.clicked.connect(self.export_all_assignments_pdf)
        
        self.export_class1_pdf_btn = QPushButton("📄 الصف الأول PDF")
        self.export_class1_pdf_btn.clicked.connect(lambda: self.export_class_assignment_pdf('الصف الأول'))
        
        self.export_class2_pdf_btn = QPushButton("📄 الصف الثاني PDF")
        self.export_class2_pdf_btn.clicked.connect(lambda: self.export_class_assignment_pdf('الصف الثاني'))
        
        self.export_class3_pdf_btn = QPushButton("📄 الصف الثالث PDF")
        self.export_class3_pdf_btn.clicked.connect(lambda: self.export_class_assignment_pdf('الصف الثالث'))
        
        pdf_buttons_layout.addWidget(self.export_all_pdf_btn)
        pdf_buttons_layout.addWidget(self.export_class1_pdf_btn)
        pdf_buttons_layout.addWidget(self.export_class2_pdf_btn)
        pdf_buttons_layout.addWidget(self.export_class3_pdf_btn)
        pdf_export_layout.addLayout(pdf_buttons_layout)
        
        self.close_btn = QPushButton("إغلاق")
        self.close_btn.clicked.connect(self.accept)
        
        control_layout.addWidget(self.distribute_btn)
        control_layout.addLayout(excel_export_layout)
        control_layout.addLayout(pdf_export_layout)
        control_layout.addStretch()
        control_layout.addWidget(self.close_btn)
        
        # تبويبات النتائج
        self.results_tabs = QTabWidget()
        
        # إنشاء تبويبات للصفوف
        self.class1_tab = QWidget()
        self.class2_tab = QWidget()
        self.class3_tab = QWidget()
        
        self.setup_assignment_table(self.class1_tab, "الصف الأول")
        self.setup_assignment_table(self.class2_tab, "الصف الثاني")
        self.setup_assignment_table(self.class3_tab, "الصف الثالث")
        
        self.results_tabs.addTab(self.class1_tab, f"الصف الأول ({class1_count})")
        self.results_tabs.addTab(self.class2_tab, f"الصف الثاني ({class2_count})")
        self.results_tabs.addTab(self.class3_tab, f"الصف الثالث ({class3_count})")
        
        # إضافة جميع العناصر
        layout.addWidget(title_label)
        layout.addWidget(stats_frame)
        layout.addWidget(settings_group)
        layout.addLayout(control_layout)
        layout.addWidget(self.results_tabs)
        
        self.setLayout(layout)
        
        # التوزيع التلقائي الأولي
        self.distribute_children()
    
    def setup_assignment_table(self, tab, class_name):
        """إعداد جدول التوزيع لكل صف مع جميع بيانات الطفل"""
        layout = QVBoxLayout()
        
        # معلومات التوزيع
        info_label = QLabel()
        info_label.setObjectName(f"info_{class_name}")
        info_label.setProperty("class", "info-label")

        layout.addWidget(info_label)
        
        # جدول التوزيع مع جميع البيانات
        table = QTableWidget()
        table.setObjectName(f"table_{class_name}")
        table.setColumnCount(15)
        table.setHorizontalHeaderLabels([
            "الخادم", "الكود", "الاسم", "الصف", "المنطقة",
            "العمارة", "الشارع", "الدور", "الشقة",
            "موبيل الولد", "موبايل الأب", "موبايل الأم", "تليفون",
            "المدرسة", "ملاحظات"
        ])
        
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(True)
        
        # ضبط أبعاد الأعمدة
        table.setColumnWidth(0, 120)  # الخادم
        table.setColumnWidth(1, 80)   # الكود
        table.setColumnWidth(2, 150)  # الاسم
        table.setColumnWidth(3, 80)   # الصف
        table.setColumnWidth(4, 100)  # المنطقة
        table.setColumnWidth(5, 80)   # العمارة
        table.setColumnWidth(6, 120)  # الشارع
        table.setColumnWidth(7, 60)   # الدور
        table.setColumnWidth(8, 60)   # الشقة
        table.setColumnWidth(9, 110)  # موبيل الولد
        table.setColumnWidth(10, 110) # موبايل الأب
        table.setColumnWidth(11, 110) # موبايل الأم
        table.setColumnWidth(12, 100) # تليفون
        table.setColumnWidth(13, 120) # المدرسة
        table.setColumnWidth(14, 200) # ملاحظات
        
        layout.addWidget(table)
        tab.setLayout(layout)
    
    def distribute_children(self):
        """توزيع الأطفال الغائبين عشوائياً على الخدام"""
        try:
            # الحصول على عدد الخدام لكل صف
            servers_count = {
                'الصف الأول': self.servers_class1.value(),
                'الصف الثاني': self.servers_class2.value(), 
                'الصف الثالث': self.servers_class3.value()
            }
            
            # فرز الأطفال حسب الصف
            children_by_class = {
                'الصف الأول': [c for c in self.absent_children if c.get('class') == 'الصف الأول'],
                'الصف الثاني': [c for c in self.absent_children if c.get('class') == 'الصف الثاني'],
                'الصف الثالث': [c for c in self.absent_children if c.get('class') == 'الصف الثالث']
            }
            
            self.server_assignments = {}
            
            for class_name, children in children_by_class.items():
                num_servers = servers_count[class_name]
                if not children:
                    continue
                
                # خلط الأطفال عشوائياً
                random.shuffle(children)
                
                # توزيع الأطفال على الخدام
                assignments = {}
                for i in range(num_servers):
                    server_name = self.default_servers[i] if i < len(self.default_servers) else f"الخادم {i+1}"
                    assignments[server_name] = []
                
                # توزيع الأطفال بالتساوي
                for i, child in enumerate(children):
                    server_index = i % num_servers
                    server_name = self.default_servers[server_index] if server_index < len(self.default_servers) else f"الخادم {server_index+1}"
                    assignments[server_name].append(child)
                
                self.server_assignments[class_name] = assignments
            
            # تحديث الجداول
            self.update_tables()
            
            QMessageBox.information(self, "تم التوزيع", "تم توزيع الأطفال الغائبين على الخدام بنجاح!")
            
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء التوزيع: {str(e)}")
    
    def update_tables(self):
        """تحديث جميع الجداول بنتائج التوزيع"""
        for class_name in ['الصف الأول', 'الصف الثاني', 'الصف الثالث']:
            self.update_class_table(class_name)
    
    def update_class_table(self, class_name):
        """تحديث جدول صف معين بجميع بيانات الطفل"""
        try:
            table = self.findChild(QTableWidget, f"table_{class_name}")
            info_label = self.findChild(QLabel, f"info_{class_name}")
            
            if not table or class_name not in self.server_assignments:
                return
            
            assignments = self.server_assignments[class_name]
            
            # تحديث معلومات التوزيع
            total_children = sum(len(children) for children in assignments.values())
            num_servers = len(assignments)
            
            info_text = f"📊 توزيع {total_children} طفل على {num_servers} خدام: "
            for server, children in assignments.items():
                info_text += f"{server} ({len(children)}) | "
            
            if info_label:
                info_label.setText(info_text[:-3])  # إزالة آخر " | "
            
            # تحديث الجدول
            table.setSortingEnabled(False)
            table.setRowCount(0)
            
            row = 0
            for server, children in assignments.items():
                for child in children:
                    table.insertRow(row)
                    
                    # الخادم
                    server_item = QTableWidgetItem(server)
                    server_item.setBackground(self.get_server_color(server))
                    table.setItem(row, 0, server_item)
                    
                    # بيانات الطفل الأساسية
                    table.setItem(row, 1, QTableWidgetItem(str(child.get('code', 'غير محدد'))))
                    table.setItem(row, 2, QTableWidgetItem(child.get('name', 'غير محدد')))
                    table.setItem(row, 3, QTableWidgetItem(child.get('class', 'غير محدد')))
                    table.setItem(row, 4, QTableWidgetItem(child.get('region', 'غير محدد')))
                    
                    # بيانات العنوان
                    table.setItem(row, 5, QTableWidgetItem(str(child.get('عماره', 'غير محدد'))))
                    table.setItem(row, 6, QTableWidgetItem(child.get('شارع', 'غير محدد')))
                    table.setItem(row, 7, QTableWidgetItem(str(child.get('دور', 'غير محدد'))))
                    table.setItem(row, 8, QTableWidgetItem(str(child.get('شقه', 'غير محدد'))))
                    
                    # بيانات الهواتف
                    table.setItem(row, 9, QTableWidgetItem(child.get('موبيل الولد', 'غير محدد')))
                    table.setItem(row, 10, QTableWidgetItem(child.get('موبايل الاب', 'غير محدد')))
                    table.setItem(row, 11, QTableWidgetItem(child.get('موبايل الام', 'غير محدد')))
                    table.setItem(row, 12, QTableWidgetItem(child.get('تليفون', 'غير محدد')))
                    
                    # بيانات إضافية
                    table.setItem(row, 13, QTableWidgetItem(child.get('المدرسه', 'غير محدد')))
                    table.setItem(row, 14, QTableWidgetItem(child.get('ملاحظات', 'غير محدد')))
                    
                    row += 1
            
            table.setSortingEnabled(True)
            
        except Exception as e:
            print(f"Error updating table for {class_name}: {str(e)}")
    
    def get_server_color(self, server_name):
        """الحصول على لون مميز لكل خادم"""
        colors = {
            'الخادم الأول': Qt.cyan,
            'الخادم الثاني': Qt.green,
            'الخادم الثالث': Qt.yellow,
            'الخادم الرابع': Qt.magenta,
            'الخادم الخامس': Qt.lightGray,
            'الخادم السادس': Qt.blue,
            'الخادم السابع': Qt.darkGreen,
            'الخادم الثامن': Qt.darkYellow,
            'الخادم التاسع': Qt.darkCyan,
            'الخادم العاشر': Qt.darkMagenta
        }
        return colors.get(server_name, Qt.white)
    
    def update_distribution(self):
        """تحديث التوزيع عند تغيير عدد الخدام"""
        if self.server_assignments:
            self.distribute_children()
    
    def export_all_assignments_excel(self):
        """تصدير جميع نتائج التوزيع إلى Excel مع تنسيق محسن"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            if not self.server_assignments:
                QMessageBox.warning(self, "تحذير", "لا توجد بيانات للتصدير. يرجى عمل توزيع أولاً.")
                return
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ تقرير التوزيع الكامل", 
                f"توزيع_المتابعة_الكامل_{self.date}.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                # إنشاء DataFrame لجميع البيانات
                all_data = []
                
                for class_name, assignments in self.server_assignments.items():
                    for server, children in assignments.items():
                        for child in children:
                            all_data.append({
                                'الصف': class_name,
                                'الخادم': server,
                                'الكود': child.get('code', 'غير محدد'),
                                'الاسم': child.get('name', 'غير محدد'),
                                'المنطقة': child.get('region', 'غير محدد'),
                                'العمارة': child.get('عماره', 'غير محدد'),
                                'الشارع': child.get('شارع', 'غير محدد'),
                                'الدور': child.get('دور', 'غير محدد'),
                                'الشقة': child.get('شقه', 'غير محدد'),
                                'موبيل الولد': child.get('موبيل الولد', 'غير محدد'),
                                'موبايل الأب': child.get('موبايل الاب', 'غير محدد'),
                                'موبايل الأم': child.get('موبايل الام', 'غير محدد'),
                                'تليفون': child.get('تليفون', 'غير محدد'),
                                'المدرسة': child.get('المدرسه', 'غير محدد'),
                                'ملاحظات': child.get('ملاحظات', 'غير محدد'),
                                'تاريخ الغياب': self.date
                            })
                
                if all_data:
                    # إنشاء ملف Excel مع تنسيق محسن
                    wb = Workbook()
                    
                    # ورقة جميع البيانات
                    ws_all = wb.active
                    ws_all.title = "جميع البيانات"
                    
                    # إضافة العناوين
                    titles = [
                        f"تقرير توزيع متابعة الغياب - الكل",
                        f"تاريخ الغياب: {self.date}",
                        f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                        f"إجمالي الأطفال الغائبين: {len(self.absent_children)}",
                        f"عدد الخدام: {sum(len(assignments) for assignments in self.server_assignments.values())}"
                    ]
                    
                    for i, title in enumerate(titles, 1):
                        ws_all[f'A{i}'] = title
                        ws_all[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                        ws_all[f'A{i}'].alignment = Alignment(horizontal='center')
                    
                    # دمج الخلايا للعناوين
                    ws_all.merge_cells('A1:O1')
                    ws_all.merge_cells('A2:O2')
                    ws_all.merge_cells('A3:O3')
                    ws_all.merge_cells('A4:O4')
                    ws_all.merge_cells('A5:O5')
                    
                    # إضافة رؤوس الأعمدة
                    headers = ['الصف', 'الخادم', 'الكود', 'الاسم', 'المنطقة', 'العمارة', 'الشارع', 'الدور', 'الشقة',
                              'موبيل الولد', 'موبايل الأب', 'موبايل الأم', 'تليفون', 'المدرسة', 'ملاحظات', 'تاريخ الغياب']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws_all.cell(row=7, column=col, value=header)
                        cell.font = Font(bold=True, color="FFFFFF", size=12)
                        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # إضافة البيانات
                    for row, data in enumerate(all_data, 8):
                        for col, key in enumerate(headers, 1):
                            cell = ws_all.cell(row=row, column=col, value=data.get(key, ''))
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                               top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # ضبط أبعاد الأعمدة تلقائياً
                    column_widths = {
                        'A': 15, 'B': 15, 'C': 12, 'D': 25, 'E': 15,
                        'F': 12, 'G': 20, 'H': 10, 'I': 10, 'J': 15,
                        'K': 15, 'L': 15, 'M': 15, 'N': 20, 'O': 30, 'P': 15
                    }
                    
                    for col, width in column_widths.items():
                        ws_all.column_dimensions[col].width = width
                    
                    # إضافة أوراق للصفوف
                    for class_name in self.server_assignments.keys():
                        ws_class = wb.create_sheet(title=class_name)
                        class_data = [d for d in all_data if d['الصف'] == class_name]
                        
                        if class_data:
                            # إضافة العناوين للصف
                            titles_class = [
                                f"تقرير توزيع متابعة الغياب - {class_name}",
                                f"تاريخ الغياب: {self.date}",
                                f"عدد الأطفال: {len(class_data)}"
                            ]
                            
                            for i, title in enumerate(titles_class, 1):
                                ws_class[f'A{i}'] = title
                                ws_class[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                                ws_class[f'A{i}'].alignment = Alignment(horizontal='center')
                            
                            ws_class.merge_cells('A1:O1')
                            ws_class.merge_cells('A2:O2')
                            ws_class.merge_cells('A3:O3')
                            
                            # رؤوس الأعمدة
                            for col, header in enumerate(headers, 1):
                                cell = ws_class.cell(row=5, column=col, value=header)
                                cell.font = Font(bold=True, color="FFFFFF", size=12)
                                cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                   top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            # البيانات
                            for row, data in enumerate(class_data, 6):
                                for col, key in enumerate(headers, 1):
                                    cell = ws_class.cell(row=row, column=col, value=data.get(key, ''))
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                       top=Side(style='thin'), bottom=Side(style='thin'))
                            
                            # ضبط الأبعاد
                            for col, width in column_widths.items():
                                ws_class.column_dimensions[col].width = width
                    
                    # حفظ الملف
                    wb.save(file_path)
                    
                    QMessageBox.information(self, "تم التصدير", 
                                          f"تم تصدير تقرير التوزيع الكامل بنجاح!\n\n"
                                          f"الملف: {file_path}\n"
                                          f"عدد الأطفال: {len(self.absent_children)}\n"
                                          f"عدد الخدام: {sum(len(assignments) for assignments in self.server_assignments.values())}")
                else:
                    QMessageBox.warning(self, "تحذير", "لا توجد بيانات للتصدير.")
                    
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء التصدير: {str(e)}")
    
    def export_class_assignment_excel(self, class_name):
        """تصدير نتائج توزيع صف معين إلى Excel مع تنسيق محسن"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            if not self.server_assignments or class_name not in self.server_assignments:
                QMessageBox.warning(self, "تحذير", f"لا توجد بيانات للصف {class_name} للتصدير.")
                return
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, f"حفظ تقرير توزيع {class_name}", 
                f"توزيع_المتابعة_{class_name}_{self.date}.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                # جمع بيانات الصف المحدد
                class_data = []
                assignments = self.server_assignments[class_name]
                
                for server, children in assignments.items():
                    for child in children:
                        class_data.append({
                            'الصف': class_name,
                            'الخادم': server,
                            'الكود': child.get('code', 'غير محدد'),
                            'الاسم': child.get('name', 'غير محدد'),
                            'المنطقة': child.get('region', 'غير محدد'),
                            'العمارة': child.get('عماره', 'غير محدد'),
                            'الشارع': child.get('شارع', 'غير محدد'),
                            'الدور': child.get('دور', 'غير محدد'),
                            'الشقة': child.get('شقه', 'غير محدد'),
                            'موبيل الولد': child.get('موبيل الولد', 'غير محدد'),
                            'موبايل الأب': child.get('موبايل الاب', 'غير محدد'),
                            'موبايل الأم': child.get('موبايل الام', 'غير محدد'),
                            'تليفون': child.get('تليفون', 'غير محدد'),
                            'المدرسة': child.get('المدرسه', 'غير محدد'),
                            'ملاحظات': child.get('ملاحظات', 'غير محدد'),
                            'تاريخ الغياب': self.date
                        })
                
                if class_data:
                    # إنشاء ملف Excel
                    wb = Workbook()
                    ws = wb.active
                    ws.title = class_name
                    
                    # إضافة العناوين
                    titles = [
                        f"تقرير توزيع متابعة الغياب - {class_name}",
                        f"تاريخ الغياب: {self.date}",
                        f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                        f"عدد الأطفال الغائبين: {len(class_data)}",
                        f"عدد الخدام: {len(assignments)}"
                    ]
                    
                    for i, title in enumerate(titles, 1):
                        ws[f'A{i}'] = title
                        ws[f'A{i}'].font = Font(bold=True, size=14 if i == 1 else 12)
                        ws[f'A{i}'].alignment = Alignment(horizontal='center')
                    
                    # دمج الخلايا
                    ws.merge_cells('A1:P1')
                    ws.merge_cells('A2:P2')
                    ws.merge_cells('A3:P3')
                    ws.merge_cells('A4:P4')
                    ws.merge_cells('A5:P5')
                    
                    # رؤوس الأعمدة
                    headers = ['الصف', 'الخادم', 'الكود', 'الاسم', 'المنطقة', 'العمارة', 'الشارع', 'الدور', 'الشقة',
                              'موبيل الولد', 'موبايل الأب', 'موبايل الأم', 'تليفون', 'المدرسة', 'ملاحظات', 'تاريخ الغياب']
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=7, column=col, value=header)
                        cell.font = Font(bold=True, color="FFFFFF", size=12)
                        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # إضافة البيانات
                    for row, data in enumerate(class_data, 8):
                        for col, key in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=data.get(key, ''))
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                               top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # ضبط أبعاد الأعمدة
                    column_widths = {
                        'A': 15, 'B': 15, 'C': 12, 'D': 25, 'E': 15,
                        'F': 12, 'G': 20, 'H': 10, 'I': 10, 'J': 15,
                        'K': 15, 'L': 15, 'M': 15, 'N': 20, 'O': 30, 'P': 15
                    }
                    
                    for col, width in column_widths.items():
                        ws.column_dimensions[col].width = width
                    
                    # حفظ الملف
                    wb.save(file_path)
                    
                    QMessageBox.information(self, "تم التصدير", 
                                          f"تم تصدير تقرير توزيع {class_name} بنجاح!\n\n"
                                          f"الملف: {file_path}\n"
                                          f"عدد الأطفال: {len(class_data)}\n"
                                          f"عدد الخدام: {len(assignments)}")
                else:
                    QMessageBox.warning(self, "تحذير", f"لا توجد بيانات للصف {class_name} للتصدير.")
                    
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء تصدير {class_name}: {str(e)}")

    def export_all_assignments_pdf(self):
        """تصدير جميع نتائج التوزيع إلى PDF مع إصلاح مشاكل العربية والتخطيط"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import arabic_reshaper
            from bidi.algorithm import get_display
            import os
            
            if not self.server_assignments:
                QMessageBox.warning(self, "تحذير", "لا توجد بيانات للتصدير. يرجى عمل توزيع أولاً.")
                return
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ تقرير PDF الكامل", 
                f"توزيع_المتابعة_الكامل_{self.date}.pdf", 
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
            
            # البحث عن خطوط عربية
            arabic_font_paths = [
                "fonts/arial.ttf",
                "fonts/tahoma.ttf",
                "fonts/arabic_font.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/tahoma.ttf"
            ]
            
            arabic_font_name = None
            for font_path in arabic_font_paths:
                if os.path.exists(font_path):
                    try:
                        font_name = os.path.splitext(os.path.basename(font_path))[0]
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                        arabic_font_name = font_name
                        break
                    except:
                        continue
            
            if not arabic_font_name:
                # استخدام خط افتراضي إذا لم يتم العثور على خط عربي
                arabic_font_name = "Helvetica"
            
            # إنشاء مستند PDF
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4), 
                                  rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()
            
            # إضافة عنوان رئيسي
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=arabic_font_name,
                fontSize=16,
                spaceAfter=20,
                alignment=1,  # مركز
                textColor=colors.HexColor('#2C3E50')
            )
            
            title_text = self.reshape_arabic("تقرير توزيع متابعة الغياب - الكل")
            elements.append(Paragraph(title_text, title_style))
            
            # معلومات التقرير
            info_style = ParagraphStyle(
                'CustomInfo',
                parent=styles['Normal'],
                fontName=arabic_font_name,
                fontSize=10,
                spaceAfter=8,
                alignment=1
            )
            
            info_texts = [
                f"تاريخ الغياب: {self.date}",
                f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                f"إجمالي الأطفال الغائبين: {len(self.absent_children)}",
                f"عدد الخدام: {sum(len(assignments) for assignments in self.server_assignments.values())}"
            ]
            
            for text in info_texts:
                elements.append(Paragraph(self.reshape_arabic(text), info_style))
            
            elements.append(Spacer(1, 20))
            
            # جمع جميع البيانات
            all_data = []
            for class_name, assignments in self.server_assignments.items():
                for server, children in assignments.items():
                    for child in children:
                        all_data.append({
                            'الصف': class_name,
                            'الخادم': server,
                            'الكود': child.get('code', 'غير محدد'),
                            'الاسم': child.get('name', 'غير محدد'),
                            'المنطقة': child.get('region', 'غير محدد'),
                            'العمارة': child.get('عماره', 'غير محدد'),
                            'الشارع': child.get('شارع', 'غير محدد'),
                            'الدور': child.get('دور', 'غير محدد'),
                            'الشقة': child.get('شقه', 'غير محدد'),
                            'موبيل الولد': child.get('موبيل الولد', 'غير محدد'),
                            'موبايل الأب': child.get('موبايل الاب', 'غير محدد'),
                            'موبايل الأم': child.get('موبايل الام', 'غير محدد'),
                            'تليفون': child.get('تليفون', 'غير محدد'),
                            'المدرسة': child.get('المدرسه', 'غير محدد'),
                            'ملاحظات': child.get('ملاحظات', 'غير محدد')
                        })
            
            if all_data:
                # رؤوس الأعمدة (من اليمين لليسار)
                headers = [
                    'ملاحظات', 'المدرسة', 'تليفون', 'موبايل الأم', 'موبايل الأب', 
                    'موبيل الولد', 'الشقة', 'الدور', 'الشارع', 'العمارة', 
                    'المنطقة', 'الصف', 'الاسم', 'الكود', 'الخادم'
                ]
                
                # إعداد بيانات الجدول
                table_data = [headers]
                
                for data in all_data:
                    row = [
                        self.reshape_arabic(data['ملاحظات']),
                        self.reshape_arabic(data['المدرسة']),
                        self.reshape_arabic(data['تليفون']),
                        self.reshape_arabic(data['موبايل الأم']),
                        self.reshape_arabic(data['موبايل الأب']),
                        self.reshape_arabic(data['موبيل الولد']),
                        self.reshape_arabic(data['الشقة']),
                        self.reshape_arabic(data['الدور']),
                        self.reshape_arabic(data['الشارع']),
                        self.reshape_arabic(data['العمارة']),
                        self.reshape_arabic(data['المنطقة']),
                        self.reshape_arabic(data['الصف']),
                        self.reshape_arabic(data['الاسم']),
                        self.reshape_arabic(data['الكود']),
                        self.reshape_arabic(data['الخادم'])
                    ]
                    table_data.append(row)
                
                # إنشاء الجدول مع أبعاد مناسبة
                table = Table(table_data, repeatRows=1)
                
                # تنسيق الجدول
                table.setStyle(TableStyle([
                    # تنسيق رأس الجدول
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), arabic_font_name),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    
                    # تنسيق بيانات الجدول
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), arabic_font_name),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    
                    # الحدود
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                    
                    # التفاف النص
                    ('WORDWRAP', (0, 0), (-1, -1), True),
                ]))
                
                # ضبط أبعاد الأعمدة
                col_widths = [80, 60, 60, 60, 60, 60, 40, 40, 60, 50, 40, 60, 50, 40, 50]
                table._argW = col_widths
                
                elements.append(table)
                
                # بناء PDF
                doc.build(elements)
                QMessageBox.information(self, "تم التصدير", "تم تصدير تقرير PDF الكامل بنجاح!")
            else:
                QMessageBox.warning(self, "تحذير", "لا توجد بيانات للتصدير.")
                
        except ImportError as e:
            QMessageBox.critical(self, "خطأ", 
                               f"المكتبات المطلوبة غير مثبتة!\n"
                               f"يرجى تثبيت المكتبات التالية:\n"
                               f"pip install reportlab arabic-reshaper python-bidi")
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء التصدير: {str(e)}")
    
    def export_class_assignment_pdf(self, class_name):
        """تصدير نتائج توزيع صف معين إلى PDF مع إصلاح مشاكل العربية"""
        try:
            from PyQt5.QtWidgets import QFileDialog
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import arabic_reshaper
            from bidi.algorithm import get_display
            import os
            
            if not self.server_assignments or class_name not in self.server_assignments:
                QMessageBox.warning(self, "تحذير", f"لا توجد بيانات للصف {class_name} للتصدير.")
                return
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, f"حفظ تقرير PDF {class_name}", 
                f"توزيع_المتابعة_{class_name}_{self.date}.pdf", 
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
            
            # البحث عن خطوط عربية
            arabic_font_paths = [
                "fonts/arial.ttf",
                "fonts/tahoma.ttf",
                "fonts/arabic_font.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/tahoma.ttf"
            ]
            
            arabic_font_name = None
            for font_path in arabic_font_paths:
                if os.path.exists(font_path):
                    try:
                        font_name = os.path.splitext(os.path.basename(font_path))[0]
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                        arabic_font_name = font_name
                        break
                    except:
                        continue
            
            if not arabic_font_name:
                arabic_font_name = "Helvetica"
            
            # جمع بيانات الصف
            class_data = []
            assignments = self.server_assignments[class_name]
            
            for server, children in assignments.items():
                for child in children:
                    class_data.append({
                        'الخادم': server,
                        'الكود': child.get('code', 'غير محدد'),
                        'الاسم': child.get('name', 'غير محدد'),
                        'الصف': class_name,
                        'المنطقة': child.get('region', 'غير محدد'),
                        'العمارة': child.get('عماره', 'غير محدد'),
                        'الشارع': child.get('شارع', 'غير محدد'),
                        'الدور': child.get('دور', 'غير محدد'),
                        'الشقة': child.get('شقه', 'غير محدد'),
                        'موبيل الولد': child.get('موبيل الولد', 'غير محدد'),
                        'موبايل الأب': child.get('موبايل الاب', 'غير محدد'),
                        'موبايل الأم': child.get('موبايل الام', 'غير محدد'),
                        'تليفون': child.get('تليفون', 'غير محدد'),
                        'المدرسة': child.get('المدرسه', 'غير محدد'),
                        'ملاحظات': child.get('ملاحظات', 'غير محدد')
                    })
            
            if class_data:
                # إنشاء مستند PDF
                doc = SimpleDocTemplate(file_path, pagesize=landscape(A4),
                                      rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
                elements = []
                styles = getSampleStyleSheet()
                
                # إضافة عنوان
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontName=arabic_font_name,
                    fontSize=16,
                    spaceAfter=20,
                    alignment=1,
                    textColor=colors.HexColor('#2C3E50')
                )
                
                title_text = self.reshape_arabic(f"تقرير توزيع متابعة الغياب - {class_name}")
                elements.append(Paragraph(title_text, title_style))
                
                # معلومات التقرير
                info_style = ParagraphStyle(
                    'CustomInfo',
                    parent=styles['Normal'],
                    fontName=arabic_font_name,
                    fontSize=10,
                    spaceAfter=8,
                    alignment=1
                )
                
                info_texts = [
                    f"تاريخ الغياب: {self.date}",
                    f"وقت التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    f"عدد الأطفال: {len(class_data)}",
                    f"عدد الخدام: {len(assignments)}"
                ]
                
                for text in info_texts:
                    elements.append(Paragraph(self.reshape_arabic(text), info_style))
                
                elements.append(Spacer(1, 20))
                
                # رؤوس الأعمدة (من اليمين لليسار)
                headers = [
                    'ملاحظات', 'المدرسة', 'تليفون', 'موبايل الأم', 'موبايل الأب', 
                    'موبيل الولد', 'الشقة', 'الدور', 'الشارع', 'العمارة', 
                    'المنطقة', 'الصف', 'الاسم', 'الكود', 'الخادم'
                ]
                
                # إعداد بيانات الجدول
                table_data = [headers]
                
                for data in class_data:
                    row = [
                        self.reshape_arabic(data['ملاحظات']),
                        self.reshape_arabic(data['المدرسة']),
                        self.reshape_arabic(data['تليفون']),
                        self.reshape_arabic(data['موبايل الأم']),
                        self.reshape_arabic(data['موبايل الأب']),
                        self.reshape_arabic(data['موبيل الولد']),
                        self.reshape_arabic(data['الشقة']),
                        self.reshape_arabic(data['الدور']),
                        self.reshape_arabic(data['الشارع']),
                        self.reshape_arabic(data['العمارة']),
                        self.reshape_arabic(data['المنطقة']),
                        self.reshape_arabic(data['الصف']),
                        self.reshape_arabic(data['الاسم']),
                        self.reshape_arabic(data['الكود']),
                        self.reshape_arabic(data['الخادم'])
                    ]
                    table_data.append(row)
                
                # إنشاء الجدول
                table = Table(table_data, repeatRows=1)
                
                # تنسيق الجدول
                table.setStyle(TableStyle([
                    # تنسيق رأس الجدول
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), arabic_font_name),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    
                    # تنسيق بيانات الجدول
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), arabic_font_name),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    
                    # الحدود
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                    
                    # التفاف النص
                    ('WORDWRAP', (0, 0), (-1, -1), True),
                ]))
                
                # ضبط أبعاد الأعمدة
                col_widths = [80, 60, 60, 60, 60, 60, 40, 40, 60, 50, 40, 60, 50, 40, 50]
                table._argW = col_widths
                
                elements.append(table)
                
                # بناء PDF
                doc.build(elements)
                QMessageBox.information(self, "تم التصدير", f"تم تصدير تقرير PDF {class_name} بنجاح!")
            else:
                QMessageBox.warning(self, "تحذير", f"لا توجد بيانات للصف {class_name} للتصدير.")
                
        except ImportError as e:
            QMessageBox.critical(self, "خطأ", 
                               f"المكتبات المطلوبة غير مثبتة!\n"
                               f"يرجى تثبيت المكتبات التالية:\n"
                               f"pip install reportlab arabic-reshaper python-bidi")
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء التصدير: {str(e)}")
    
    def reshape_arabic(self, text):
        """إعادة تشكيل النص العربي للعرض الصحيح في PDF"""
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            
            if not text or text == 'غير محدد':
                return 'غير محدد'
            
            # إعادة تشكيل النص العربي
            reshaped_text = arabic_reshaper.reshape(str(text))
            # تحويل النص إلى تنسيق من اليمين لليسار
            bidi_text = get_display(reshaped_text)
            return bidi_text
        except:
            # في حالة حدوث خطأ، إرجاع النص الأصلي
            return str(text)